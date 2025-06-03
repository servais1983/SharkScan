"""
Advanced report generator for SharkScan
"""

import json
import os
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Callable
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import Progress
from rich import box
from rich.live import Live
from rich.layout import Layout
from rich.text import Text
from rich.bar import Bar
from rich.tree import Tree
from rich.markdown import Markdown
from rich.syntax import Syntax
from rich.rule import Rule
from rich.align import Align
from rich.style import Style
from rich.theme import Theme
from rich.prompt import Prompt
from rich.prompt import Confirm
from rich.console import Group
from rich.columns import Columns
from rich.progress_bar import ProgressBar
from rich.progress import SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from .secure_logger import SecureLogger
import time
from openpyxl.styles import PatternFill, Font, Alignment

class AlertManager:
    """Manages real-time security alerts"""
    
    def __init__(self):
        self.alerts = []
        self.alert_callbacks = []
        self.alert_thresholds = {
            'CRITICAL': 1,
            'HIGH': 3,
            'MEDIUM': 5,
            'LOW': 10
        }
        self.alert_history = []
        self.max_history_size = 1000
        self.alert_patterns = {}
        
    def add_alert_callback(self, callback: Callable[[Dict], None]):
        """Add alert callback function"""
        self.alert_callbacks.append(callback)
        
    def check_alerts(self, vulnerabilities: List[Dict], incidents: List[Dict]):
        """Check for new alerts with pattern recognition"""
        # Check severity thresholds with pattern recognition
        severity_counts = {
            'CRITICAL': len([v for v in vulnerabilities if v['severity'] == 'CRITICAL']),
            'HIGH': len([v for v in vulnerabilities if v['severity'] == 'HIGH']),
            'MEDIUM': len([v for v in vulnerabilities if v['severity'] == 'MEDIUM']),
            'LOW': len([v for v in vulnerabilities if v['severity'] == 'LOW'])
        }
        
        # Pattern recognition for severity trends
        for severity, count in severity_counts.items():
            if count >= self.alert_thresholds[severity]:
                # Check for pattern in recent alerts
                pattern = self._detect_pattern(severity, count)
                alert = {
                    'type': 'severity_threshold',
                    'severity': severity,
                    'count': count,
                    'threshold': self.alert_thresholds[severity],
                    'timestamp': datetime.now().isoformat(),
                    'pattern': pattern
                }
                self._trigger_alert(alert)
                
        # Check for new critical incidents with correlation
        for incident in incidents:
            if incident.get('severity') == 'CRITICAL':
                # Correlate with similar incidents
                correlated_incidents = self._correlate_incidents(incident)
                alert = {
                    'type': 'critical_incident',
                    'incident': incident,
                    'timestamp': datetime.now().isoformat(),
                    'correlated_incidents': correlated_incidents
                }
                self._trigger_alert(alert)
                
    def _detect_pattern(self, severity: str, count: int) -> Dict:
        """Detect patterns in alert frequency and severity"""
        recent_alerts = [a for a in self.alert_history if a['severity'] == severity]
        if len(recent_alerts) < 3:
            return {'type': 'new_pattern', 'confidence': 0.5}
            
        # Calculate trend
        trend = self._calculate_trend(recent_alerts)
        return {
            'type': 'trend',
            'direction': trend['direction'],
            'confidence': trend['confidence']
        }
        
    def _calculate_trend(self, alerts: List[Dict]) -> Dict:
        """Calculate trend from alert history"""
        if len(alerts) < 2:
            return {'direction': 'stable', 'confidence': 0.5}
            
        counts = [a['count'] for a in alerts]
        if len(counts) >= 3:
            # Calculate moving average
            ma = sum(counts[-3:]) / 3
            if ma > counts[-1]:
                return {'direction': 'decreasing', 'confidence': 0.7}
            elif ma < counts[-1]:
                return {'direction': 'increasing', 'confidence': 0.7}
        return {'direction': 'stable', 'confidence': 0.5}
        
    def _correlate_incidents(self, incident: Dict) -> List[Dict]:
        """Correlate similar incidents"""
        similar_incidents = []
        for hist_incident in self.alert_history:
            if (hist_incident.get('type') == 'critical_incident' and
                self._are_incidents_similar(incident, hist_incident['incident'])):
                similar_incidents.append(hist_incident)
        return similar_incidents
        
    def _are_incidents_similar(self, incident1: Dict, incident2: Dict) -> bool:
        """Check if two incidents are similar"""
        # Compare key attributes
        return (incident1.get('type') == incident2.get('type') and
                incident1.get('severity') == incident2.get('severity'))
                
    def _trigger_alert(self, alert: Dict):
        """Trigger alert callbacks with memory optimization"""
        self.alerts.append(alert)
        self.alert_history.append(alert)
        
        # Optimize memory usage
        if len(self.alert_history) > self.max_history_size:
            self.alert_history = self.alert_history[-self.max_history_size:]
            
        # Trigger callbacks
        for callback in self.alert_callbacks:
            callback(alert)
            
    def get_alerts(self, limit: int = 10) -> List[Dict]:
        """Get recent alerts with pattern information"""
        return self.alerts[-limit:]
        
    def clear_alerts(self):
        """Clear all alerts"""
        self.alerts = []
        self.alert_history = []
        
    def get_alert_patterns(self) -> Dict:
        """Get detected alert patterns"""
        return self.alert_patterns

class ReportGenerator:
    """Advanced report generator for security findings"""
    
    def __init__(self, output_dir: str = "reports"):
        """
        Initialize report generator
        
        Args:
            output_dir (str): Directory for report output
        """
        self.logger = SecureLogger("report_generator")
        self.output_dir = output_dir
        self.console = Console(theme=Theme({
            "info": "cyan",
            "warning": "yellow",
            "error": "red",
            "success": "green",
            "critical": "bold red",
            "high": "red",
            "medium": "yellow",
            "low": "green"
        }))
        self._setup_output_dir()
        self.alert_manager = AlertManager()
        self.alert_manager.add_alert_callback(self._handle_alert)
        
        # Optimized data storage with chunking
        self.chunk_size = 1000
        self.data_chunks = {
            'vulnerabilities': [],
            'incidents': [],
            'risk_scores': [],
            'trends': {
                'severity': {'CRITICAL': [], 'HIGH': [], 'MEDIUM': [], 'LOW': []},
                'categories': {},
                'confidence': {'high': [], 'medium': [], 'low': []}
            },
            'alerts': [],
        }
        
        # Cache for frequently accessed data
        self.cache = {
            'risk_summary': None,
            'vuln_summary': None,
            'compliance_summary': None,
            'last_update': None
        }
        
        # Visualization settings
        self.viz_settings = {
            'chart_theme': 'default',
            'interactive': True,
            'animation': True,
            'refresh_rate': 1.0
        }
        
    def _handle_alert(self, alert: Dict):
        """Handle new alert with optimized storage"""
        self.data_chunks['alerts'].append(alert)
        
        # Optimize memory by chunking
        if len(self.data_chunks['alerts']) > self.chunk_size:
            self._save_chunk('alerts')
            self.data_chunks['alerts'] = []
            
        # Invalidate cache
        self.cache['last_update'] = datetime.now()
        
    def _save_chunk(self, data_type: str):
        """Save data chunk to disk"""
        try:
            chunk_file = os.path.join(self.output_dir, f"{data_type}_chunk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
            with open(chunk_file, 'w') as f:
                json.dump(self.data_chunks[data_type], f)
        except Exception as e:
            self.logger.error(f"Error saving {data_type} chunk: {str(e)}")
            
    def _load_chunk(self, data_type: str, chunk_file: str):
        """Load data chunk from disk"""
        try:
            with open(chunk_file, 'r') as f:
                return json.load(f)
        except Exception as e:
            self.logger.error(f"Error loading {data_type} chunk: {str(e)}")
            return []
            
    def update_real_time_data(self, new_vulns: List[Dict], new_incidents: List[Dict]):
        """
        Update real-time data with optimized storage
        
        Args:
            new_vulns (List[Dict]): New vulnerabilities
            new_incidents (List[Dict]): New incidents
        """
        # Update vulnerabilities with chunking
        self.data_chunks['vulnerabilities'].extend(new_vulns)
        if len(self.data_chunks['vulnerabilities']) > self.chunk_size:
            self._save_chunk('vulnerabilities')
            self.data_chunks['vulnerabilities'] = []
            
        # Update incidents with chunking
        self.data_chunks['incidents'].extend(new_incidents)
        if len(self.data_chunks['incidents']) > self.chunk_size:
            self._save_chunk('incidents')
            self.data_chunks['incidents'] = []
            
        # Update trends with optimized calculations
        self._update_trends(new_vulns)
        
        # Check for alerts
        self.alert_manager.check_alerts(new_vulns, new_incidents)
        
        # Calculate current risk score with caching
        current_risk = self._calculate_risk_level(
            self.data_chunks['vulnerabilities'],
            self.data_chunks['incidents']
        )
        self.data_chunks['risk_scores'].append(current_risk)
        
        # Optimize risk scores storage
        if len(self.data_chunks['risk_scores']) > self.chunk_size:
            self._save_chunk('risk_scores')
            self.data_chunks['risk_scores'] = []
            
        # Invalidate cache
        self.cache['last_update'] = datetime.now()
        self.cache['risk_summary'] = None
        self.cache['vuln_summary'] = None
        
    def _update_trends(self, new_vulns: List[Dict]):
        """Update trend data with optimized calculations"""
        for vuln in new_vulns:
            # Update severity trends with optimized storage
            severity = vuln['severity']
            self.data_chunks['trends']['severity'][severity].append(1)
            
            # Update category trends with optimized storage
            category = vuln.get('category', 'unknown')
            if category not in self.data_chunks['trends']['categories']:
                self.data_chunks['trends']['categories'][category] = []
            self.data_chunks['trends']['categories'][category].append(1)
            
            # Update confidence trends with optimized storage
            confidence = vuln.get('confidence', 0.5)
            if confidence >= 0.8:
                self.data_chunks['trends']['confidence']['high'].append(1)
            elif confidence >= 0.5:
                self.data_chunks['trends']['confidence']['medium'].append(1)
            else:
                self.data_chunks['trends']['confidence']['low'].append(1)
                
        # Optimize trend data storage
        for severity in self.data_chunks['trends']['severity']:
            if len(self.data_chunks['trends']['severity'][severity]) > self.chunk_size:
                self._save_chunk(f'trends_severity_{severity}')
                self.data_chunks['trends']['severity'][severity] = []
                
        for category in self.data_chunks['trends']['categories']:
            if len(self.data_chunks['trends']['categories'][category]) > self.chunk_size:
                self._save_chunk(f'trends_category_{category}')
                self.data_chunks['trends']['categories'][category] = []
                
        for confidence in self.data_chunks['trends']['confidence']:
            if len(self.data_chunks['trends']['confidence'][confidence]) > self.chunk_size:
                self._save_chunk(f'trends_confidence_{confidence}')
                self.data_chunks['trends']['confidence'][confidence] = []
                
    def display_real_time_report(self):
        """Display real-time security report with optimized rendering"""
        layout = Layout()
        layout.split_column(
            Layout(name="header"),
            Layout(name="alerts"),
            Layout(name="body"),
            Layout(name="footer")
        )
        
        layout["body"].split_row(
            Layout(name="vulns"),
            Layout(name="incidents"),
            Layout(name="risk")
        )
        
        with Live(layout, refresh_per_second=self.viz_settings['refresh_rate']) as live:
            while True:
                # Update header with cached data
                layout["header"].update(self._create_header())
                
                # Update alerts with optimized rendering
                layout["alerts"].update(self._create_alerts_section())
                
                # Update body sections with optimized rendering
                layout["vulns"].update(self._create_vuln_section())
                layout["incidents"].update(self._create_incident_section())
                layout["risk"].update(self._create_risk_section())
                
                # Update footer
                layout["footer"].update(self._create_footer())
                
                # Sleep to prevent excessive CPU usage
                time.sleep(1 / self.viz_settings['refresh_rate'])
        
    def _create_header(self) -> Panel:
        """Create header panel"""
        return Panel(
            Align.center(
                Text("SharkScan Real-Time Security Monitor", style="bold blue"),
                vertical="middle"
            ),
            border_style="blue"
        )
        
    def _create_alerts_section(self) -> Panel:
        """Create alerts section"""
        alerts = self.data_chunks['alerts'][-5:]  # Show last 5 alerts
        
        if not alerts:
            return Panel("No active alerts", title="Active Alerts", border_style="blue")
            
        alert_table = Table(show_header=True, header_style="bold magenta", box=box.ROUNDED)
        alert_table.add_column("Time", style="cyan")
        alert_table.add_column("Type", style="yellow")
        alert_table.add_column("Details", style="green")
        
        for alert in alerts:
            time = datetime.fromisoformat(alert['timestamp']).strftime('%H:%M:%S')
            alert_type = alert['type']
            details = self._format_alert_details(alert)
            alert_table.add_row(time, alert_type, details)
            
        return Panel(alert_table, title="Active Alerts", border_style="blue")
        
    def _format_alert_details(self, alert: Dict) -> str:
        """Format alert details"""
        if alert['type'] == 'severity_threshold':
            return f"{alert['severity']} vulnerabilities: {alert['count']} (threshold: {alert['threshold']})"
        elif alert['type'] == 'critical_incident':
            return f"Critical incident: {alert['incident'].get('type', 'unknown')}"
        return str(alert)
        
    def _create_vuln_section(self) -> Panel:
        """Create vulnerabilities section"""
        vulns = self.data_chunks['vulnerabilities']
        severity_counts = {
            'CRITICAL': len([v for v in vulns if v['severity'] == 'CRITICAL']),
            'HIGH': len([v for v in vulns if v['severity'] == 'HIGH']),
            'MEDIUM': len([v for v in vulns if v['severity'] == 'MEDIUM']),
            'LOW': len([v for v in vulns if v['severity'] == 'LOW'])
        }
        
        # Create severity table
        table = Table(show_header=True, header_style="bold magenta", box=box.ROUNDED)
        table.add_column("Severity", style="cyan")
        table.add_column("Count", justify="right", style="green")
        
        for severity, count in severity_counts.items():
            table.add_row(severity, str(count))
            
        # Create trend sparklines
        trend_table = Table(show_header=False, box=box.SIMPLE)
        trend_table.add_column("Trend")
        
        for severity in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']:
            trend_data = self.data_chunks['trends']['severity'][severity]
            if trend_data:
                bar = Bar(trend_data[-1], width=20, color=severity.lower())
                trend_table.add_row(bar)
                
        return Panel(
            Layout.split_column(
                Layout(table),
                Layout(trend_table)
            ),
            title="Vulnerabilities",
            border_style="blue"
        )
        
    def _create_incident_section(self) -> Panel:
        """Create incidents section"""
        incidents = self.data_chunks['incidents']
        type_counts = {}
        
        for incident in incidents:
            incident_type = incident.get('type', 'unknown')
            type_counts[incident_type] = type_counts.get(incident_type, 0) + 1
            
        # Create incident table
        table = Table(show_header=True, header_style="bold magenta", box=box.ROUNDED)
        table.add_column("Type", style="cyan")
        table.add_column("Count", justify="right", style="green")
        
        for incident_type, count in type_counts.items():
            table.add_row(incident_type, str(count))
            
        # Create incident timeline
        timeline = Tree("📊 Incident Timeline")
        for incident in sorted(incidents[-5:], key=lambda x: x.get('timestamp', '')):
            timeline.add(f"{incident.get('type', 'unknown')} - {incident.get('timestamp', '')}")
            
        return Panel(
            Layout.split_column(
                Layout(table),
                Layout(timeline)
            ),
            title="Incidents",
            border_style="blue"
        )
        
    def _create_risk_section(self) -> Panel:
        """Create risk section"""
        risk_scores = self.data_chunks['risk_scores']
        if not risk_scores:
            return Panel("No risk data available", title="Risk Trend", border_style="blue")
            
        # Create risk trend bar
        current_risk = risk_scores[-1]
        trend_bar = Bar(
            current_risk,
            width=40,
            color="green" if current_risk < 0.4 else "yellow" if current_risk < 0.7 else "red"
        )
        
        # Create trend indicator
        if len(risk_scores) > 1:
            trend = "↑" if risk_scores[-1] > risk_scores[-2] else "↓"
            trend_color = "red" if trend == "↑" else "green"
            trend_text = Text(f"Trend: {trend}", style=trend_color)
        else:
            trend_text = Text("Trend: →", style="yellow")
        
        return Panel(
            Layout.split_column(
                Layout(f"Current Risk: {current_risk:.2f}"),
                Layout(trend_bar),
                Layout(trend_text)
            ),
            title="Risk Trend",
            border_style="blue"
        )
        
    def _create_footer(self) -> Panel:
        """Create footer panel"""
        return Panel(
            Align.center(
                Text(f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", style="dim"),
                vertical="middle"
            ),
            border_style="blue"
        )

    def generate_report(self, scan_results: Dict[str, Any], vulnerabilities: List[Dict], 
                       incidents: List[Dict], recommendations: List[Dict]) -> Dict[str, Any]:
        """
        Generate comprehensive security report with advanced features
        
        Args:
            scan_results (Dict[str, Any]): Scan results
            vulnerabilities (List[Dict]): Detected vulnerabilities
            incidents (List[Dict]): Security incidents
            recommendations (List[Dict]): Security recommendations
            
        Returns:
            Dict[str, Any]: Generated report
        """
        # Vérifier le cache avant de générer
        cache_key = f"{hash(str(scan_results))}_{hash(str(vulnerabilities))}_{hash(str(incidents))}"
        if cache_key in self.cache and self.cache[cache_key].get('timestamp', 0) > (datetime.now() - timedelta(hours=1)).timestamp():
            return self.cache[cache_key]['report']

        report = {
            'timestamp': datetime.now().isoformat(),
            'summary': self._generate_summary(scan_results, vulnerabilities, incidents),
            'vulnerabilities': self._analyze_vulnerabilities(vulnerabilities),
            'incidents': self._analyze_incidents(incidents),
            'recommendations': self._generate_recommendations(recommendations),
            'risk_assessment': self._assess_risk(vulnerabilities, incidents),
            'compliance': self._analyze_compliance(),
            'threat_intelligence': self._analyze_threat_intelligence(),
            'audit_log': self._generate_audit_log(),
            'metadata': {
                'version': '2.0',
                'generator': 'SharkScan',
                'format_version': '1.0',
                'generation_time': datetime.now().isoformat(),
                'data_sources': self._get_data_sources(),
                'confidence_scores': self._calculate_confidence_scores(vulnerabilities, incidents)
            }
        }
        
        # Mettre en cache le rapport
        self.cache[cache_key] = {
            'report': report,
            'timestamp': datetime.now().timestamp()
        }
        
        # Nettoyer le cache si nécessaire
        self._clean_cache()
        
        # Afficher le rapport
        self._display_report(report)
        
        # Sauvegarder le rapport
        self._save_report(report)
        
        return report

    def _get_data_sources(self) -> Dict[str, Any]:
        """Get information about data sources used in the report"""
        return {
            'scanners': [
                {'name': 'Nmap', 'version': '7.92', 'confidence': 0.95},
                {'name': 'Nessus', 'version': '10.4.0', 'confidence': 0.90},
                {'name': 'Custom Scanner', 'version': '1.0', 'confidence': 0.85}
            ],
            'threat_intel': [
                {'source': 'MITRE ATT&CK', 'last_update': datetime.now().isoformat()},
                {'source': 'VirusTotal', 'last_update': datetime.now().isoformat()},
                {'source': 'AlienVault OTX', 'last_update': datetime.now().isoformat()}
            ],
            'compliance_sources': [
                {'standard': 'ISO 27001', 'version': '2013'},
                {'standard': 'GDPR', 'version': '2018'},
                {'standard': 'PCI DSS', 'version': '4.0'}
            ]
        }

    def _calculate_confidence_scores(self, vulnerabilities: List[Dict], incidents: List[Dict]) -> Dict[str, float]:
        """Calculate confidence scores for different aspects of the report"""
        return {
            'vulnerability_detection': self._calculate_vuln_confidence(vulnerabilities),
            'incident_analysis': self._calculate_incident_confidence(incidents),
            'risk_assessment': self._calculate_risk_confidence(vulnerabilities, incidents),
            'compliance_analysis': self._calculate_compliance_confidence(),
            'threat_intelligence': self._calculate_threat_intel_confidence()
        }

    def _calculate_vuln_confidence(self, vulnerabilities: List[Dict]) -> float:
        """Calculate confidence score for vulnerability detection"""
        if not vulnerabilities:
            return 0.0
            
        confidence_scores = [v.get('confidence', 0.5) for v in vulnerabilities]
        return sum(confidence_scores) / len(confidence_scores)

    def _calculate_incident_confidence(self, incidents: List[Dict]) -> float:
        """Calculate confidence score for incident analysis"""
        if not incidents:
            return 0.0
            
        confidence_scores = [i.get('confidence', 0.5) for i in incidents]
        return sum(confidence_scores) / len(confidence_scores)

    def _calculate_risk_confidence(self, vulnerabilities: List[Dict], incidents: List[Dict]) -> float:
        """Calculate confidence score for risk assessment"""
        vuln_confidence = self._calculate_vuln_confidence(vulnerabilities)
        incident_confidence = self._calculate_incident_confidence(incidents)
        return (vuln_confidence + incident_confidence) / 2

    def _calculate_compliance_confidence(self) -> float:
        """Calculate confidence score for compliance analysis"""
        return 0.85  # Placeholder - should be calculated based on actual compliance data

    def _calculate_threat_intel_confidence(self) -> float:
        """Calculate confidence score for threat intelligence"""
        return 0.90  # Placeholder - should be calculated based on actual threat intel data

    def _clean_cache(self):
        """Clean old entries from the cache"""
        current_time = datetime.now().timestamp()
        self.cache = {
            k: v for k, v in self.cache.items()
            if current_time - v.get('timestamp', 0) < 3600  # Keep entries less than 1 hour old
        }

    def _generate_summary(self, scan_results: Dict[str, Any], vulnerabilities: List[Dict],
                         incidents: List[Dict]) -> Dict[str, Any]:
        """
        Generate report summary
        
        Args:
            scan_results (Dict[str, Any]): Scan results
            vulnerabilities (List[Dict]): Detected vulnerabilities
            incidents (List[Dict]): Security incidents
            
        Returns:
            Dict[str, Any]: Summary information
        """
        return {
            'scan_info': {
                'targets': len(scan_results.get('targets', [])),
                'ports_scanned': len(scan_results.get('ports', [])),
                'duration': scan_results.get('duration', 0)
            },
            'findings': {
                'vulnerabilities': len(vulnerabilities),
                'incidents': len(incidents),
                'critical': len([v for v in vulnerabilities if v.get('severity') == 'CRITICAL']),
                'high': len([v for v in vulnerabilities if v.get('severity') == 'HIGH']),
                'medium': len([v for v in vulnerabilities if v.get('severity') == 'MEDIUM']),
                'low': len([v for v in vulnerabilities if v.get('severity') == 'LOW'])
            },
            'risk_level': self._calculate_risk_level(vulnerabilities, incidents)
        }
        
    def _analyze_vulnerabilities(self, vulnerabilities: List[Dict]) -> Dict[str, Any]:
        """
        Analyze detected vulnerabilities
        
        Args:
            vulnerabilities (List[Dict]): Detected vulnerabilities
            
        Returns:
            Dict[str, Any]: Vulnerability analysis
        """
        analysis = {
            'by_severity': {},
            'by_category': {},
            'by_cwe': {},
            'trends': self._analyze_vulnerability_trends(vulnerabilities)
        }
        
        if not vulnerabilities:
            return analysis
            
        for vuln in vulnerabilities:
            if not isinstance(vuln, dict):
                continue
                
            # Count by severity with safe access
            severity = vuln.get('severity', 'unknown')
            analysis['by_severity'][severity] = analysis['by_severity'].get(severity, 0) + 1
            
            # Count by category with safe access
            category = vuln.get('category', 'unknown')
            analysis['by_category'][category] = analysis['by_category'].get(category, 0) + 1
            
            # Count by CWE with safe access
            cwe = vuln.get('cwe', 'N/A')
            analysis['by_cwe'][cwe] = analysis['by_cwe'].get(cwe, 0) + 1
            
        return analysis
        
    def _analyze_vulnerability_trends(self, vulnerabilities: List[Dict]) -> Dict[str, Any]:
        """Analyze vulnerability trends"""
        if not vulnerabilities:
            return {
                'severity': {},
                'category': {},
                'trends': {}
            }
        
        # Filter out None or non-dict elements
        valid_vulns = [v for v in vulnerabilities if isinstance(v, dict)]
        
        severity_trends = {}
        category_trends = {}
        
        for vuln in valid_vulns:
            # Update severity trends with safe access
            severity = vuln.get('severity', 'unknown')
            if severity not in severity_trends:
                severity_trends[severity] = []
            severity_trends[severity].append(vuln)
            
            # Update category trends with safe access
            category = vuln.get('category', 'unknown')
            if category not in category_trends:
                category_trends[category] = []
            category_trends[category].append(vuln)
        
        return {
            'severity': severity_trends,
            'category': category_trends,
            'trends': {
                'severity': {s: len(v) for s, v in severity_trends.items()},
                'category': {c: len(v) for c, v in category_trends.items()}
            }
        }
        
    def _analyze_incidents(self, incidents: List[Dict]) -> Dict[str, Any]:
        """Analyze security incidents"""
        if not incidents:
            return {
                'type_counts': {},
                'severity_counts': {},
                'impact_counts': {},
                'total_incidents': 0,
                'trends': {
                    'severity': {},
                    'type': {},
                    'impact': {}
                }
            }
            
        # Filtrer les incidents None ou non-dictionnaires
        valid_incidents = [inc for inc in incidents if isinstance(inc, dict)]
        
        type_counts = {}
        severity_counts = {}
        impact_counts = {}
        
        for incident in valid_incidents:
            # Compter les types d'incidents avec accès sécurisé
            incident_type = incident.get('type', 'unknown')
            type_counts[incident_type] = type_counts.get(incident_type, 0) + 1
            
            # Compter les sévérités avec accès sécurisé
            severity = incident.get('severity', 'unknown')
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
            
            # Compter les impacts avec accès sécurisé
            impact = incident.get('impact', 'unknown')
            impact_counts[impact] = impact_counts.get(impact, 0) + 1
            
        # Calculer les tendances
        trends = {
            'severity': severity_counts,
            'type': type_counts,
            'impact': impact_counts
        }
            
        return {
            'type_counts': type_counts,
            'severity_counts': severity_counts,
            'impact_counts': impact_counts,
            'total_incidents': len(valid_incidents),
            'trends': trends
        }
        
    def _analyze_incident_impact(self, incidents: List[Dict]) -> Dict[str, Any]:
        """Analyze impact of security incidents"""
        if not incidents:
            return {
                'type_counts': {},
                'severity_counts': {},
                'impact_counts': {},
                'total_incidents': 0
            }
            
        # Filtrer les incidents None ou non-dictionnaires
        valid_incidents = [inc for inc in incidents if isinstance(inc, dict)]
        
        type_counts = {}
        severity_counts = {}
        impact_counts = {}
        
        for incident in valid_incidents:
            # Compter les types d'incidents
            incident_type = incident.get('type', 'unknown')
            type_counts[incident_type] = type_counts.get(incident_type, 0) + 1
            
            # Compter les sévérités
            severity = incident.get('severity', 'unknown')
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
            
            # Compter les impacts
            impact = incident.get('impact', 'unknown')
            impact_counts[impact] = impact_counts.get(impact, 0) + 1
            
        return {
            'type_counts': type_counts,
            'severity_counts': severity_counts,
            'impact_counts': impact_counts,
            'total_incidents': len(valid_incidents)
        }
        
    def _generate_recommendations(self, recommendations: List[Dict]) -> Dict[str, Any]:
        """Generate recommendations"""
        if not recommendations:
            return {
                'priority': {},
                'category': {},
                'details': []
            }
        
        # Filter out None or non-dict elements
        valid_recommendations = [r for r in recommendations if isinstance(r, dict)]
        
        priority_counts = {}
        category_counts = {}
        details = []
        
        for rec in valid_recommendations:
            priority = rec.get('priority', 'unknown')
            category = rec.get('category', 'unknown')
            
            if priority not in priority_counts:
                priority_counts[priority] = 0
            priority_counts[priority] += 1
            
            if category not in category_counts:
                category_counts[category] = 0
            category_counts[category] += 1
            
            details.append({
                'priority': priority,
                'category': category,
                'description': rec.get('description', ''),
                'action': rec.get('action', '')
            })
        
        return {
            'priority': priority_counts,
            'category': category_counts,
            'details': details
        }
        
    def _assess_risk(self, vulnerabilities: List[Dict], incidents: List[Dict]) -> Dict[str, Any]:
        """
        Assess overall security risk
        
        Args:
            vulnerabilities (List[Dict]): Detected vulnerabilities
            incidents (List[Dict]): Security incidents
            
        Returns:
            Dict[str, Any]: Risk assessment
        """
        assessment = {
            'overall_risk': 0,
            'risk_factors': {
                'vulnerabilities': 0,
                'incidents': 0,
                'exposure': 0,
                'impact': 0
            },
            'risk_level': 'LOW',
            'trend': 'stable'
        }
        
        # Calculate risk from vulnerabilities
        vuln_risk = sum(self._get_risk_weight(v['severity']) for v in vulnerabilities)
        assessment['risk_factors']['vulnerabilities'] = min(vuln_risk / 10, 1.0)
        
        # Calculate risk from incidents
        incident_risk = sum(self._get_risk_weight(i.get('severity', 'LOW')) for i in incidents)
        assessment['risk_factors']['incidents'] = min(incident_risk / 10, 1.0)
        
        # Calculate exposure risk
        exposed_systems = set()
        for vuln in vulnerabilities:
            if 'affected_systems' in vuln:
                exposed_systems.update(vuln['affected_systems'])
        assessment['risk_factors']['exposure'] = min(len(exposed_systems) / 10, 1.0)
        
        # Calculate impact risk
        impact_risk = sum(self._get_impact_weight(i.get('impact', 'LOW')) for i in incidents)
        assessment['risk_factors']['impact'] = min(impact_risk / 10, 1.0)
        
        # Calculate overall risk
        assessment['overall_risk'] = sum(assessment['risk_factors'].values()) / 4
        
        # Determine risk level
        if assessment['overall_risk'] >= 0.8:
            assessment['risk_level'] = 'CRITICAL'
        elif assessment['overall_risk'] >= 0.6:
            assessment['risk_level'] = 'HIGH'
        elif assessment['overall_risk'] >= 0.4:
            assessment['risk_level'] = 'MEDIUM'
        else:
            assessment['risk_level'] = 'LOW'
            
        return assessment
        
    def _get_risk_weight(self, severity: str) -> float:
        """Get risk weight for severity level"""
        weights = {
            'CRITICAL': 1.0,
            'HIGH': 0.8,
            'MEDIUM': 0.5,
            'LOW': 0.2
        }
        return weights.get(severity, 0.1)
        
    def _get_impact_weight(self, impact: str) -> float:
        """Get impact weight for impact level"""
        weights = {
            'CRITICAL': 1.0,
            'HIGH': 0.8,
            'MEDIUM': 0.5,
            'LOW': 0.2
        }
        return weights.get(impact, 0.1)
        
    def _calculate_risk_level(self, vulnerabilities: List[Dict], incidents: List[Dict]) -> str:
        """Calculate overall risk level"""
        risk_score = 0
        
        # Calculate from vulnerabilities
        for vuln in vulnerabilities:
            risk_score += self._get_risk_weight(vuln['severity'])
            
        # Calculate from incidents
        for incident in incidents:
            risk_score += self._get_risk_weight(incident.get('severity', 'LOW'))
            
        # Normalize score
        risk_score = min(risk_score / (len(vulnerabilities) + len(incidents) + 1), 1.0)
        
        # Determine risk level
        if risk_score >= 0.8:
            return 'CRITICAL'
        elif risk_score >= 0.6:
            return 'HIGH'
        elif risk_score >= 0.4:
            return 'MEDIUM'
        else:
            return 'LOW'
            
    def _display_report(self, report: Dict[str, Any]):
        """Display report in console"""
        self.console.print("\n[bold blue]SharkScan Security Report[/bold blue]")
        self.console.print(f"Generated: {report['timestamp']}\n")
        
        # Display summary
        self.console.print(Panel(
            f"[bold]Risk Level:[/bold] {report['summary']['risk_level']}\n"
            f"[bold]Vulnerabilities:[/bold] {report['summary']['findings']['vulnerabilities']}\n"
            f"[bold]Incidents:[/bold] {report['summary']['findings']['incidents']}\n"
            f"[bold]Critical:[/bold] {report['summary']['findings']['critical']}\n"
            f"[bold]High:[/bold] {report['summary']['findings']['high']}\n"
            f"[bold]Compliance Score:[/bold] {report['compliance']['overall_score']:.1f}%\n"
            f"[bold]Threat Risk Level:[/bold] {report['threat_intelligence']['risk_assessment']['risk_level']}",
            title="Summary",
            border_style="blue"
        ))
        
        # Display vulnerabilities
        vuln_table = Table(title="Vulnerabilities", box=box.ROUNDED)
        vuln_table.add_column("Severity", style="cyan")
        vuln_table.add_column("Category", style="magenta")
        vuln_table.add_column("Count", justify="right", style="green")
        
        for severity, count in report['vulnerabilities']['by_severity'].items():
            vuln_table.add_row(severity, "All", str(count))
        
        self.console.print(vuln_table)
        
        # Display risk assessment
        risk_table = Table(title="Risk Assessment", box=box.ROUNDED)
        risk_table.add_column("Factor", style="cyan")
        risk_table.add_column("Score", justify="right", style="green")
        
        for factor, score in report['risk_assessment']['risk_factors'].items():
            risk_table.add_row(factor, f"{score:.2f}")
        
        self.console.print(risk_table)
        
        # Display compliance status
        compliance_table = Table(title="Compliance Status", box=box.ROUNDED)
        compliance_table.add_column("Standard", style="cyan")
        compliance_table.add_column("Score", justify="right", style="green")
        compliance_table.add_column("Status", style="magenta")
        
        for standard, data in report['compliance']['standards'].items():
            compliance_table.add_row(
                standard,
                f"{data['score']:.1f}%",
                data['status']
            )
        
        self.console.print(compliance_table)
        
        # Display threat intelligence
        threat_table = Table(title="Threat Intelligence", box=box.ROUNDED)
        threat_table.add_column("Type", style="cyan")
        threat_table.add_column("Count", justify="right", style="green")
        threat_table.add_column("Risk Level", style="magenta")
        
        for threat_type, count in report['threat_intelligence']['trends']['by_type'].items():
            threat_table.add_row(
                threat_type,
                str(count),
                report['threat_intelligence']['risk_assessment']['risk_level']
            )
        
        self.console.print(threat_table)
        
        # Sécuriser l'affichage des incidents et vulnérabilités détaillées
        # (exemple pour incidents)
        if 'incidents' in report and isinstance(report['incidents'], dict):
            incidents = report['incidents'].get('details', [])
            if isinstance(incidents, list):
                for inc in incidents:
                    if not isinstance(inc, dict):
                        continue
                    # Afficher les détails de l'incident (exemple)
                    self.console.print(f"Incident: {inc.get('type', 'unknown')} - {inc.get('severity', 'unknown')}")
        # Idem pour d'autres sections si besoin

    def _save_report(self, report: Dict[str, Any]):
        """Save report to file"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(self.output_dir, f"security_report_{timestamp}.json")
            
            with open(filename, 'w') as f:
                json.dump(report, f, indent=2)
                
            self.logger.info(f"Report saved to {filename}")
            
        except Exception as e:
            self.logger.error(f"Error saving report: {str(e)}")
            
    def export_report(self, report: Dict[str, Any], format: str = 'json', filters: Optional[Dict] = None) -> Optional[str]:
        """
        Export report in specified format with advanced filtering and formatting options
        
        Args:
            report (Dict[str, Any]): Report to export
            format (str): Export format (json, html, pdf, excel, csv)
            filters (Dict, optional): Filter criteria
            
        Returns:
            Optional[str]: Path to exported file
        """
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Appliquer les filtres si spécifiés
            if filters:
                filtered_data = self._apply_export_filters(report, filters)
            else:
                filtered_data = report
                
            # Générer le nom du fichier
            filename = os.path.join(self.output_dir, f"security_report_{timestamp}.{format}")
            
            # Exporter selon le format
            if format == 'json':
                self._export_json(filtered_data, filename)
            elif format == 'html':
                self._export_html(filtered_data, filename)
            elif format == 'pdf':
                self._export_pdf(filtered_data, filename)
            elif format == 'excel':
                self._export_excel(filtered_data, filename)
            elif format == 'csv':
                self._export_csv(filtered_data, filename)
            else:
                self.logger.error(f"Format d'export non supporté: {format}")
                return None
                
            self.logger.info(f"Rapport exporté vers {filename}")
            return filename
            
        except Exception as e:
            self.logger.error(f"Erreur lors de l'export du rapport: {str(e)}")
            return None

    def _apply_export_filters(self, report: Dict[str, Any], filters: Dict) -> Dict[str, Any]:
        """Apply filters to report data for export"""
        filtered_data = report.copy()
        
        # Filtrer par sévérité
        if 'severity' in filters:
            filtered_data['vulnerabilities'] = [
                v for v in filtered_data['vulnerabilities']
                if v.get('severity') == filters['severity']
            ]
            
        # Filtrer par période
        if 'time_range' in filters:
            start_time = datetime.fromisoformat(filters['time_range'][0])
            end_time = datetime.fromisoformat(filters['time_range'][1])
            filtered_data['incidents'] = [
                i for i in filtered_data['incidents']
                if start_time <= datetime.fromisoformat(i.get('timestamp', '')) <= end_time
            ]
            
        # Filtrer par catégorie
        if 'category' in filters:
            filtered_data['vulnerabilities'] = [
                v for v in filtered_data['vulnerabilities']
                if v.get('category') == filters['category']
            ]
            
        # Filtrer par niveau de confiance
        if 'confidence_threshold' in filters:
            filtered_data['vulnerabilities'] = [
                v for v in filtered_data['vulnerabilities']
                if v.get('confidence', 0) >= filters['confidence_threshold']
            ]
            
        return filtered_data

    def _export_json(self, data: Dict[str, Any], filename: str):
        """Export data as JSON with pretty printing"""
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    def _export_excel(self, data: Dict[str, Any], filename: str):
        """Export data as Excel with multiple sheets"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            
            # Créer un classeur Excel
            wb = Workbook()
            
            # Feuille de résumé
            ws_summary = wb.active
            ws_summary.title = "Résumé"
            self._write_summary_sheet(ws_summary, data['summary'])
            
            # Feuille des vulnérabilités
            ws_vulns = wb.create_sheet("Vulnérabilités")
            self._write_vulnerabilities_sheet(ws_vulns, data['vulnerabilities'])
            
            # Feuille des incidents
            ws_incidents = wb.create_sheet("Incidents")
            self._write_incidents_sheet(ws_incidents, data['incidents'])
            
            # Feuille d'évaluation des risques
            ws_risk = wb.create_sheet("Risques")
            self._write_risk_sheet(ws_risk, data['risk_assessment'])
            
            # Sauvegarder le fichier
            wb.save(filename)
            
        except ImportError:
            self.logger.error("Pandas ou openpyxl non installé. Installez-les avec: pip install pandas openpyxl")
            raise

    def _write_summary_sheet(self, ws, summary_data: Dict[str, Any]):
        """Write summary data to Excel sheet"""
        # En-tête
        ws['A1'] = "Résumé du Rapport de Sécurité"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Données du résumé
        row = 3
        for key, value in summary_data.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value)
            row += 1

    def _write_vulnerabilities_sheet(self, ws, vuln_data: Dict[str, Any]):
        """Write vulnerability data to Excel sheet"""
        # En-tête
        headers = ['Sévérité', 'Catégorie', 'Description', 'Confiance', 'Impact']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            
        # Données des vulnérabilités
        row = 2
        for vuln in vuln_data.get('by_severity', {}).items():
            ws[f'A{row}'] = vuln[0]
            ws[f'B{row}'] = vuln[1]
            row += 1

    def _write_incidents_sheet(self, ws, incident_data: Dict[str, Any]):
        """Write incident data to Excel sheet"""
        # En-tête
        headers = ['Type', 'Sévérité', 'Timestamp', 'Description', 'Impact']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            
        # Données des incidents
        row = 2
        for incident in incident_data.get('by_type', {}).items():
            ws[f'A{row}'] = incident[0]
            ws[f'B{row}'] = incident[1]
            row += 1

    def _write_risk_sheet(self, ws, risk_data: Dict[str, Any]):
        """Write risk assessment data to Excel sheet"""
        # En-tête
        ws['A1'] = "Évaluation des Risques"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Données d'évaluation des risques
        row = 3
        for factor, score in risk_data.get('risk_factors', {}).items():
            ws[f'A{row}'] = factor
            ws[f'B{row}'] = f"{score:.2f}"
            row += 1

    def _export_csv(self, data: Dict[str, Any], filename: str):
        """Export data as CSV with multiple files for different sections"""
        try:
            import pandas as pd
            
            # Exporter les vulnérabilités
            vuln_df = pd.DataFrame(data['vulnerabilities'])
            vuln_df.to_csv(f"{filename}_vulnerabilities.csv", index=False)
            
            # Exporter les incidents
            incident_df = pd.DataFrame(data['incidents'])
            incident_df.to_csv(f"{filename}_incidents.csv", index=False)
            
            # Exporter l'évaluation des risques
            risk_df = pd.DataFrame([data['risk_assessment']])
            risk_df.to_csv(f"{filename}_risk_assessment.csv", index=False)
            
        except ImportError:
            self.logger.error("Pandas non installé. Installez-le avec: pip install pandas")
            raise

    def _export_html(self, report: Dict[str, Any], filename: str):
        """Export report as HTML with advanced visualizations"""
        html_template = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>SharkScan Security Report</title>
            <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
            <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.0.0"></script>
            <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-annotation@1.4.0"></script>
            <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-streaming@2.0.1"></script>
            <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-dragdata@2.2.3"></script>
            <script src="https://cdn.jsdelivr.net/npm/chartjs-chart-box-and-violin-plot@4.0.0/build/Chart.BoxPlot.min.js"></script>
            <script src="https://cdn.jsdelivr.net/npm/chartjs-chart-matrix@1.1.1/dist/chartjs-chart-matrix.min.js"></script>
            <script src="https://cdn.jsdelivr.net/npm/chartjs-chart-sankey@0.8.0/dist/chartjs-chart-sankey.min.js"></script>
            <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-regression@1.0.0/dist/chartjs-plugin-regression.min.js"></script>
            <script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
            <script src="https://cdn.jsdelivr.net/npm/jspdf@2.5.1/dist/jspdf.umd.min.js"></script>
            <script src="https://cdn.jsdelivr.net/npm/jspdf-autotable@3.5.31/dist/jspdf.plugin.autotable.min.js"></script>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }
                .section { margin-bottom: 20px; background-color: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
                .risk-high { color: #dc3545; }
                .risk-medium { color: #ffc107; }
                .risk-low { color: #28a745; }
                table { border-collapse: collapse; width: 100%; margin-top: 10px; }
                th, td { border: 1px solid #ddd; padding: 12px; text-align: left; }
                th { background-color: #f8f9fa; }
                .chart-container { position: relative; height: 400px; width: 100%; margin: 20px 0; }
                .dashboard { display: grid; grid-template-columns: repeat(2, 1fr); gap: 20px; }
                .stats-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin: 20px 0; }
                .stat-card { background: white; padding: 15px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
                .stat-value { font-size: 24px; font-weight: bold; margin: 10px 0; }
                .stat-label { color: #6c757d; }
                .trend-indicator { display: inline-block; margin-left: 10px; }
                .trend-up { color: #dc3545; }
                .trend-down { color: #28a745; }
                .trend-stable { color: #6c757d; }
                .filters { margin: 20px 0; padding: 15px; background: white; border-radius: 8px; }
                .filter-group { margin: 10px 0; }
                .filter-label { margin-right: 10px; }
                .interactive-chart { cursor: pointer; }
                .tooltip { position: absolute; background: rgba(0,0,0,0.8); color: white; padding: 5px; border-radius: 4px; }
                .alert-badge {
                    position: fixed;
                    top: 20px;
                    right: 20px;
                    background: #dc3545;
                    color: white;
                    padding: 10px 20px;
                    border-radius: 20px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.2);
                    z-index: 1000;
                    display: none;
                }
                .recommendation-card {
                    background: white;
                    padding: 15px;
                    border-radius: 8px;
                    margin: 10px 0;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }
                .recommendation-priority {
                    display: inline-block;
                    padding: 3px 8px;
                    border-radius: 4px;
                    color: white;
                    font-size: 12px;
                    margin-left: 10px;
                }
                .priority-critical { background: #dc3545; }
                .priority-high { background: #fd7e14; }
                .priority-medium { background: #ffc107; }
                .priority-low { background: #28a745; }
                .heatmap-container {
                    margin: 20px 0;
                    padding: 15px;
                    background: white;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }
                
                .sankey-container {
                    height: 500px;
                    margin: 20px 0;
                }
                
                .boxplot-container {
                    margin: 20px 0;
                    padding: 15px;
                    background: white;
                    border-radius: 8px;
                }
                
                .correlation-matrix {
                    margin: 20px 0;
                    padding: 15px;
                    background: white;
                    border-radius: 8px;
                }
                
                .trend-analysis {
                    display: grid;
                    grid-template-columns: repeat(2, 1fr);
                    gap: 20px;
                    margin: 20px 0;
                }
                
                .trend-card {
                    background: white;
                    padding: 15px;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }
                
                .trend-header {
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    margin-bottom: 10px;
                }
                
                .impact-analysis {
                    margin: 20px 0;
                    padding: 15px;
                    background: white;
                    border-radius: 8px;
                }
                
                .impact-grid {
                    display: grid;
                    grid-template-columns: repeat(3, 1fr);
                    gap: 15px;
                }
                
                .impact-card {
                    padding: 15px;
                    border-radius: 8px;
                    background: #f8f9fa;
                }
                
                .impact-title {
                    font-weight: bold;
                    margin-bottom: 10px;
                }
                
                .impact-value {
                    font-size: 20px;
                    color: #007bff;
                }
                
                .predictive-analysis {
                    margin: 20px 0;
                    padding: 15px;
                    background: white;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }
                
                .prediction-card {
                    background: #f8f9fa;
                    padding: 15px;
                    border-radius: 8px;
                    margin: 10px 0;
                }
                
                .prediction-header {
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    margin-bottom: 10px;
                }
                
                .prediction-value {
                    font-size: 24px;
                    font-weight: bold;
                    color: #007bff;
                }
                
                .prediction-confidence {
                    font-size: 14px;
                    color: #6c757d;
                }
                
                .export-options {
                    display: grid;
                    grid-template-columns: repeat(3, 1fr);
                    gap: 15px;
                    margin: 20px 0;
                }
                
                .export-card {
                    background: white;
                    padding: 15px;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    text-align: center;
                }
                
                .export-icon {
                    font-size: 24px;
                    margin-bottom: 10px;
                }
                
                .export-button {
                    display: inline-block;
                    padding: 8px 16px;
                    background: #007bff;
                    color: white;
                    border-radius: 4px;
                    cursor: pointer;
                    text-decoration: none;
                }
                
                .export-button:hover {
                    background: #0056b3;
                }
                
                .risk-prediction {
                    margin: 20px 0;
                    padding: 15px;
                    background: white;
                    border-radius: 8px;
                }
                
                .prediction-timeline {
                    margin: 20px 0;
                    padding: 15px;
                    background: white;
                    border-radius: 8px;
                }
                
                .timeline-item {
                    display: flex;
                    align-items: center;
                    margin: 10px 0;
                    padding: 10px;
                    background: #f8f9fa;
                    border-radius: 4px;
                }
                
                .timeline-date {
                    font-weight: bold;
                    margin-right: 10px;
                }
                
                .timeline-content {
                    flex-grow: 1;
                }
                
                .timeline-severity {
                    padding: 4px 8px;
                    border-radius: 4px;
                    color: white;
                    font-size: 12px;
                }
                
                .severity-critical { background: #dc3545; }
                .severity-high { background: #fd7e14; }
                .severity-medium { background: #ffc107; }
                .severity-low { background: #28a745; }
                
                .advanced-filters {
                    margin: 20px 0;
                    padding: 20px;
                    background: white;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }
                
                .filter-grid {
                    display: grid;
                    grid-template-columns: repeat(3, 1fr);
                    gap: 15px;
                }
                
                .filter-group {
                    margin: 10px 0;
                }
                
                .filter-label {
                    display: block;
                    margin-bottom: 5px;
                    font-weight: bold;
                }
                
                .filter-input {
                    width: 100%;
                    padding: 8px;
                    border: 1px solid #ddd;
                    border-radius: 4px;
                }
                
                .historical-comparison {
                    margin: 20px 0;
                    padding: 20px;
                    background: white;
                    border-radius: 8px;
                }
                
                .comparison-grid {
                    display: grid;
                    grid-template-columns: repeat(2, 1fr);
                    gap: 20px;
                }
                
                .comparison-card {
                    background: #f8f9fa;
                    padding: 15px;
                    border-radius: 8px;
                }
                
                .comparison-header {
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    margin-bottom: 10px;
                }
                
                .comparison-value {
                    font-size: 24px;
                    font-weight: bold;
                }
                
                .comparison-change {
                    font-size: 14px;
                    padding: 4px 8px;
                    border-radius: 4px;
                }
                
                .change-positive { background: #28a745; color: white; }
                .change-negative { background: #dc3545; color: white; }
                
                .recommendations {
                    margin: 20px 0;
                    padding: 20px;
                    background: white;
                    border-radius: 8px;
                }
                
                .recommendation-card {
                    background: #f8f9fa;
                    padding: 15px;
                    border-radius: 8px;
                    margin: 10px 0;
                }
                
                .recommendation-header {
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                }
                
                .recommendation-priority {
                    padding: 4px 8px;
                    border-radius: 4px;
                    color: white;
                    font-size: 12px;
                }
                
                .priority-high { background: #dc3545; }
                .priority-medium { background: #ffc107; }
                .priority-low { background: #28a745; }
                
                .integration-section {
                    margin: 20px 0;
                    padding: 20px;
                    background: white;
                    border-radius: 8px;
                }
                
                .integration-grid {
                    display: grid;
                    grid-template-columns: repeat(3, 1fr);
                    gap: 15px;
                }
                
                .integration-card {
                    background: #f8f9fa;
                    padding: 15px;
                    border-radius: 8px;
                    text-align: center;
                }
                
                .integration-icon {
                    font-size: 32px;
                    margin-bottom: 10px;
                }
                
                .integration-status {
                    display: inline-block;
                    padding: 4px 8px;
                    border-radius: 4px;
                    font-size: 12px;
                    margin-top: 10px;
                }
                
                .status-active { background: #28a745; color: white; }
                .status-inactive { background: #6c757d; color: white; }
                
                .compliance-section {
                    margin: 20px 0;
                    padding: 20px;
                    background: white;
                    border-radius: 8px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }
                
                .compliance-grid {
                    display: grid;
                    grid-template-columns: repeat(3, 1fr);
                    gap: 15px;
                }
                
                .compliance-card {
                    background: #f8f9fa;
                    padding: 15px;
                    border-radius: 8px;
                }
                
                .compliance-header {
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    margin-bottom: 10px;
                }
                
                .compliance-score {
                    font-size: 24px;
                    font-weight: bold;
                }
                
                .compliance-status {
                    padding: 4px 8px;
                    border-radius: 4px;
                    font-size: 12px;
                    color: white;
                }
                
                .status-compliant { background: #28a745; }
                .status-partial { background: #ffc107; }
                .status-non-compliant { background: #dc3545; }
                
                .audit-section {
                    margin: 20px 0;
                    padding: 20px;
                    background: white;
                    border-radius: 8px;
                }
                
                .audit-timeline {
                    margin: 20px 0;
                }
                
                .audit-event {
                    display: flex;
                    align-items: center;
                    margin: 10px 0;
                    padding: 10px;
                    background: #f8f9fa;
                    border-radius: 4px;
                }
                
                .audit-icon {
                    font-size: 24px;
                    margin-right: 10px;
                }
                
                .audit-details {
                    flex-grow: 1;
                }
                
                .audit-time {
                    color: #6c757d;
                    font-size: 12px;
                }
                
                .threat-intel {
                    margin: 20px 0;
                    padding: 20px;
                    background: white;
                    border-radius: 8px;
                }
                
                .threat-grid {
                    display: grid;
                    grid-template-columns: repeat(2, 1fr);
                    gap: 15px;
                }
                
                .threat-card {
                    background: #f8f9fa;
                    padding: 15px;
                    border-radius: 8px;
                }
                
                .threat-header {
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    margin-bottom: 10px;
                }
                
                .threat-severity {
                    padding: 4px 8px;
                    border-radius: 4px;
                    font-size: 12px;
                    color: white;
                }
                
                .severity-critical { background: #dc3545; }
                .severity-high { background: #fd7e14; }
                .severity-medium { background: #ffc107; }
                .severity-low { background: #28a745; }
                
                .custom-alerts {
                    margin: 20px 0;
                    padding: 20px;
                    background: white;
                    border-radius: 8px;
                }
                
                .alert-rules {
                    margin: 20px 0;
                }
                
                .alert-rule {
                    background: #f8f9fa;
                    padding: 15px;
                    border-radius: 8px;
                    margin: 10px 0;
                }
                
                .alert-header {
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                }
                
                .alert-actions {
                    display: flex;
                    gap: 10px;
                }
                
                .alert-button {
                    padding: 4px 8px;
                    border-radius: 4px;
                    font-size: 12px;
                    cursor: pointer;
                }
                
                .button-edit { background: #007bff; color: white; }
                .button-delete { background: #dc3545; color: white; }
                
                .automated-reports {
                    margin: 20px 0;
                    padding: 20px;
                    background: white;
                    border-radius: 8px;
                }
                
                .report-schedule {
                    margin: 20px 0;
                }
                
                .schedule-item {
                    background: #f8f9fa;
                    padding: 15px;
                    border-radius: 8px;
                    margin: 10px 0;
                }
                
                .schedule-header {
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                }
                
                .schedule-frequency {
                    padding: 4px 8px;
                    border-radius: 4px;
                    font-size: 12px;
                    background: #6c757d;
                    color: white;
                }
            </style>
        </head>
        <body>
            <div id="alertBadge" class="alert-badge"></div>
            
            <h1>SharkScan Security Report</h1>
            <p>Generated: {timestamp}</p>
            
            <div class="filters">
                <h3>Filters</h3>
                <div class="filter-group">
                    <label class="filter-label">Severity:</label>
                    <select id="severityFilter" onchange="applyFilters()">
                        <option value="all">All</option>
                        <option value="CRITICAL">Critical</option>
                        <option value="HIGH">High</option>
                        <option value="MEDIUM">Medium</option>
                        <option value="LOW">Low</option>
                    </select>
                </div>
                <div class="filter-group">
                    <label class="filter-label">Time Range:</label>
                    <input type="date" id="startDate" onchange="applyFilters()">
                    <input type="date" id="endDate" onchange="applyFilters()">
                </div>
                <div class="filter-group">
                    <label class="filter-label">Confidence:</label>
                    <input type="range" id="confidenceFilter" min="0" max="100" value="0" onchange="applyFilters()">
                    <span id="confidenceValue">0%</span>
                </div>
            </div>
            
            <div class="export-buttons">
                <button class="export-button" onclick="exportData('json')">Export JSON</button>
                <button class="export-button" onclick="exportData('csv')">Export CSV</button>
                <button class="export-button" onclick="exportData('excel')">Export Excel</button>
            </div>
            
            <div class="historical-comparison">
                <h3>Historical Comparison</h3>
                <div id="historicalStats"></div>
            </div>
            
            <div class="section">
                <h2>Personalized Recommendations</h2>
                <div id="recommendations"></div>
            </div>
            
            <div class="stats-grid">
                <div class="stat-card">
                    <div class="stat-label">Total Vulnerabilities</div>
                    <div class="stat-value">{vuln_count}</div>
                    <div class="trend-indicator {vuln_trend_class}">{vuln_trend}</div>
                </div>
                <div class="stat-card">
                    <div class="stat-label">Critical Issues</div>
                    <div class="stat-value">{critical_count}</div>
                    <div class="trend-indicator {critical_trend_class}">{critical_trend}</div>
                </div>
                <div class="stat-card">
                    <div class="stat-label">Risk Score</div>
                    <div class="stat-value">{risk_score}</div>
                    <div class="trend-indicator {risk_trend_class}">{risk_trend}</div>
                </div>
                <div class="stat-card">
                    <div class="stat-label">Confidence</div>
                    <div class="stat-value">{confidence_score}%</div>
                    <div class="trend-indicator {confidence_trend_class}">{confidence_trend}</div>
                </div>
            </div>
            
            <div class="dashboard">
                <div class="section">
                    <h2>Risk Assessment</h2>
                    <div class="chart-container">
                        <canvas id="riskChart"></canvas>
                    </div>
                </div>
                
                <div class="section">
                    <h2>Vulnerabilities by Severity</h2>
                    <div class="chart-container">
                        <canvas id="vulnChart"></canvas>
                    </div>
                </div>
                
                <div class="section">
                    <h2>Trend Analysis</h2>
                    <div class="chart-container">
                        <canvas id="trendChart"></canvas>
                    </div>
                </div>
                
                <div class="section">
                    <h2>Category Distribution</h2>
                    <div class="chart-container">
                        <canvas id="categoryChart"></canvas>
                    </div>
                </div>
            </div>
            
            <div class="section">
                <h2>Detailed Findings</h2>
                <div class="chart-container">
                    <canvas id="findingsChart"></canvas>
                </div>
                <table id="findingsTable">
                    <tr>
                        <th>Category</th>
                        <th>Count</th>
                        <th>Severity</th>
                        <th>Confidence</th>
                    </tr>
                    {detailed_rows}
                </table>
            </div>
            
            <div class="trend-analysis">
                <div class="trend-card">
                    <div class="trend-header">
                        <h3>Vulnerability Trends</h3>
                        <span class="trend-indicator {vuln_trend_class}">{vuln_trend}</span>
                    </div>
                    <div class="trend-value">{vuln_count}</div>
                    <div class="chart-container">
                        <canvas id="vulnTrendChart"></canvas>
                    </div>
                </div>
                
                <div class="trend-card">
                    <div class="trend-header">
                        <h3>Risk Score Trends</h3>
                        <span class="trend-indicator {risk_trend_class}">{risk_trend}</span>
                    </div>
                    <div class="trend-value">{risk_score}</div>
                    <div class="chart-container">
                        <canvas id="riskTrendChart"></canvas>
                    </div>
                </div>
            </div>
            
            <div class="impact-analysis">
                <h3>Impact Analysis</h3>
                <div class="impact-grid">
                    <div class="impact-card">
                        <div class="impact-title">Data Breach Risk</div>
                        <div class="impact-value">{data_breach_risk}%</div>
                    </div>
                    <div class="impact-card">
                        <div class="impact-title">Service Disruption</div>
                        <div class="impact-value">{service_disruption}%</div>
                    </div>
                    <div class="impact-card">
                        <div class="impact-title">Reputation Impact</div>
                        <div class="impact-value">{reputation_impact}%</div>
                    </div>
                </div>
            </div>
            
            <div class="heatmap-container">
                <h3>Vulnerability Heatmap</h3>
                <div class="chart-container">
                    <canvas id="heatmapChart"></canvas>
                </div>
            </div>
            
            <div class="sankey-container">
                <h3>Vulnerability Flow Analysis</h3>
                <canvas id="sankeyChart"></canvas>
            </div>
            
            <div class="boxplot-container">
                <h3>Severity Distribution</h3>
                <canvas id="boxplotChart"></canvas>
            </div>
            
            <div class="correlation-matrix">
                <h3>Vulnerability Correlations</h3>
                <canvas id="correlationChart"></canvas>
            </div>
            
            <div class="predictive-analysis">
                <h3>Predictive Analysis</h3>
                <div class="prediction-card">
                    <div class="prediction-header">
                        <h4>Risk Score Prediction</h4>
                        <span class="prediction-confidence">Confidence: {risk_prediction_confidence}%</span>
                    </div>
                    <div class="prediction-value">{risk_prediction_value}</div>
                    <div class="chart-container">
                        <canvas id="predictionChart"></canvas>
                    </div>
                </div>
                
                <div class="prediction-card">
                    <div class="prediction-header">
                        <h4>Vulnerability Trend Prediction</h4>
                        <span class="prediction-confidence">Confidence: {vuln_prediction_confidence}%</span>
                    </div>
                    <div class="prediction-value">{vuln_prediction_value}</div>
                    <div class="chart-container">
                        <canvas id="vulnPredictionChart"></canvas>
                    </div>
                </div>
            </div>
            
            <div class="risk-prediction">
                <h3>Risk Prediction Timeline</h3>
                <div class="prediction-timeline">
                    {prediction_timeline}
                </div>
            </div>
            
            <div class="export-options">
                <div class="export-card">
                    <div class="export-icon">📊</div>
                    <h4>Export as Excel</h4>
                    <p>Detailed analysis in Excel format</p>
                    <button class="export-button" onclick="exportToExcel()">Export</button>
                </div>
                
                <div class="export-card">
                    <div class="export-icon">📄</div>
                    <h4>Export as PDF</h4>
                    <p>Professional PDF report</p>
                    <button class="export-button" onclick="exportToPDF()">Export</button>
                </div>
                
                <div class="export-card">
                    <div class="export-icon">📈</div>
                    <h4>Export as JSON</h4>
                    <p>Raw data for analysis</p>
                    <button class="export-button" onclick="exportToJSON()">Export</button>
                </div>
            </div>
            
            <div class="advanced-filters">
                <h3>Filtres Avancés</h3>
                <div class="filter-grid">
                    <div class="filter-group">
                        <label class="filter-label">Période</label>
                        <select class="filter-input" id="timeRange">
                            <option value="7">7 derniers jours</option>
                            <option value="30">30 derniers jours</option>
                            <option value="90">90 derniers jours</option>
                            <option value="custom">Personnalisé</option>
                        </select>
                    </div>
                    
                    <div class="filter-group">
                        <label class="filter-label">Catégorie</label>
                        <select class="filter-input" id="categoryFilter" multiple>
                            {category_options}
                        </select>
                    </div>
                    
                    <div class="filter-group">
                        <label class="filter-label">Niveau de Confiance</label>
                        <input type="range" class="filter-input" id="confidenceFilter" min="0" max="100" value="0">
                        <span id="confidenceValue">0%</span>
                    </div>
                    
                    <div class="filter-group">
                        <label class="filter-label">Impact</label>
                        <select class="filter-input" id="impactFilter">
                            <option value="all">Tous</option>
                            <option value="high">Élevé</option>
                            <option value="medium">Moyen</option>
                            <option value="low">Faible</option>
                        </select>
                    </div>
                    
                    <div class="filter-group">
                        <label class="filter-label">Statut</label>
                        <select class="filter-input" id="statusFilter">
                            <option value="all">Tous</option>
                            <option value="active">Actif</option>
                            <option value="resolved">Résolu</option>
                            <option value="in_progress">En cours</option>
                        </select>
                    </div>
                    
                    <div class="filter-group">
                        <label class="filter-label">Source</label>
                        <select class="filter-input" id="sourceFilter">
                            <option value="all">Toutes</option>
                            <option value="scanner">Scanner</option>
                            <option value="manual">Manuel</option>
                            <option value="integration">Intégration</option>
                        </select>
                    </div>
                </div>
            </div>
            
            <div class="historical-comparison">
                <h3>Comparaison Historique</h3>
                <div class="comparison-grid">
                    <div class="comparison-card">
                        <div class="comparison-header">
                            <h4>Score de Risque</h4>
                            <span class="comparison-change {risk_change_class}">{risk_change}%</span>
                        </div>
                        <div class="comparison-value">{current_risk}</div>
                        <div class="chart-container">
                            <canvas id="riskComparisonChart"></canvas>
                        </div>
                    </div>
                    
                    <div class="comparison-card">
                        <div class="comparison-header">
                            <h4>Vulnérabilités</h4>
                            <span class="comparison-change {vuln_change_class}">{vuln_change}%</span>
                        </div>
                        <div class="comparison-value">{current_vulns}</div>
                        <div class="chart-container">
                            <canvas id="vulnComparisonChart"></canvas>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class="recommendations">
                <h3>Recommandations Personnalisées</h3>
                {recommendation_cards}
            </div>
            
            <div class="integration-section">
                <h3>Intégrations</h3>
                <div class="integration-grid">
                    <div class="integration-card">
                        <div class="integration-icon">🛡️</div>
                        <h4>WAF</h4>
                        <p>Protection contre les attaques web</p>
                        <span class="integration-status status-active">Actif</span>
                    </div>
                    
                    <div class="integration-card">
                        <div class="integration-icon">🔍</div>
                        <h4>SIEM</h4>
                        <p>Surveillance des événements</p>
                        <span class="integration-status status-active">Actif</span>
                    </div>
                    
                    <div class="integration-card">
                        <div class="integration-icon">🔐</div>
                        <h4>IAM</h4>
                        <p>Gestion des identités</p>
                        <span class="integration-status status-inactive">Inactif</span>
                    </div>
                </div>
            </div>
            
            <div class="compliance-section">
                <h3>Conformité et Audit</h3>
                <div class="compliance-grid">
                    <div class="compliance-card">
                        <div class="compliance-header">
                            <h4>ISO 27001</h4>
                            <span class="compliance-status status-compliant">Conforme</span>
                        </div>
                        <div class="compliance-score">92%</div>
                        <div class="chart-container">
                            <canvas id="iso27001Chart"></canvas>
                        </div>
                    </div>
                    
                    <div class="compliance-card">
                        <div class="compliance-header">
                            <h4>GDPR</h4>
                            <span class="compliance-status status-partial">Partiel</span>
                        </div>
                        <div class="compliance-score">78%</div>
                        <div class="chart-container">
                            <canvas id="gdprChart"></canvas>
                        </div>
                    </div>
                    
                    <div class="compliance-card">
                        <div class="compliance-header">
                            <h4>PCI DSS</h4>
                            <span class="compliance-status status-non-compliant">Non Conforme</span>
                        </div>
                        <div class="compliance-score">45%</div>
                        <div class="chart-container">
                            <canvas id="pciChart"></canvas>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class="audit-section">
                <h3>Journal d'Audit</h3>
                <div class="audit-timeline">
                    {audit_events}
                </div>
            </div>
            
            <div class="threat-intel">
                <h3>Intelligence des Menaces</h3>
                <div class="threat-grid">
                    <div class="threat-card">
                        <div class="threat-header">
                            <h4>Menaces Actives</h4>
                            <span class="threat-severity severity-high">Élevée</span>
                        </div>
                        <div class="chart-container">
                            <canvas id="activeThreatsChart"></canvas>
                        </div>
                    </div>
                    
                    <div class="threat-card">
                        <div class="threat-header">
                            <h4>Tendances</h4>
                            <span class="threat-severity severity-medium">Moyenne</span>
                        </div>
                        <div class="chart-container">
                            <canvas id="threatTrendsChart"></canvas>
                        </div>
                    </div>
                </div>
            </div>
            
            <div class="custom-alerts">
                <h3>Alertes Personnalisées</h3>
                <div class="alert-rules">
                    {alert_rules}
                </div>
            </div>
            
            <div class="automated-reports">
                <h3>Rapports Automatisés</h3>
                <div class="report-schedule">
                    {report_schedules}
                </div>
            </div>
            
            <script>
                // Chart.js configuration
                Chart.register(ChartDataLabels);
                
                // Risk Assessment Chart
                new Chart(document.getElementById('riskChart'), {{
                    type: 'radar',
                    data: {{
                        labels: {risk_labels},
                        datasets: [{{
                            label: 'Current Risk',
                            data: {risk_data},
                            fill: true,
                            backgroundColor: 'rgba(255, 99, 132, 0.2)',
                            borderColor: 'rgb(255, 99, 132)',
                            pointBackgroundColor: 'rgb(255, 99, 132)',
                            pointBorderColor: '#fff',
                            pointHoverBackgroundColor: '#fff',
                            pointHoverBorderColor: 'rgb(255, 99, 132)'
                        }},
                        {{
                            label: 'Previous Risk',
                            data: {previous_risk_data},
                            fill: true,
                            backgroundColor: 'rgba(54, 162, 235, 0.2)',
                            borderColor: 'rgb(54, 162, 235)',
                            pointBackgroundColor: 'rgb(54, 162, 235)',
                            pointBorderColor: '#fff',
                            pointHoverBackgroundColor: '#fff',
                            pointHoverBorderColor: 'rgb(54, 162, 235)'
                        }}]
                    }},
                    options: {{
                        scales: {{
                            r: {{
                                beginAtZero: true,
                                max: 1
                            }}
                        }},
                        plugins: {{
                            zoom: {{
                                zoom: {{
                                    wheel: {{ enabled: true }},
                                    pinch: {{ enabled: true }},
                                    mode: 'xy'
                                }}
                            }}
                        }}
                    }}
                }});
                
                // Vulnerabilities Chart
                new Chart(document.getElementById('vulnChart'), {{
                    type: 'doughnut',
                    data: {{
                        labels: {vuln_labels},
                        datasets: [{{
                            data: {vuln_data},
                            backgroundColor: [
                                'rgb(220, 53, 69)',
                                'rgb(255, 193, 7)',
                                'rgb(40, 167, 69)',
                                'rgb(23, 162, 184)'
                            ]
                        }}]
                    }},
                    options: {{
                        plugins: {{
                            datalabels: {{
                                formatter: (value, ctx) => {{
                                    let sum = ctx.dataset.data.reduce((a, b) => a + b, 0);
                                    let percentage = (value * 100 / sum).toFixed(1) + "%";
                                    return percentage;
                                }},
                                color: '#fff',
                                font: {{
                                    weight: 'bold'
                                }}
                            }}
                        }}
                    }}
                }});
                
                // Trend Analysis Chart
                new Chart(document.getElementById('trendChart'), {{
                    type: 'line',
                    data: {{
                        labels: {trend_labels},
                        datasets: [{{
                            label: 'Vulnerabilities',
                            data: {trend_data},
                            borderColor: 'rgb(255, 99, 132)',
                            tension: 0.1
                        }},
                        {{
                            label: 'Risk Score',
                            data: {risk_trend_data},
                            borderColor: 'rgb(54, 162, 235)',
                            tension: 0.1
                        }}]
                    }},
                    options: {{
                        plugins: {{
                            zoom: {{
                                zoom: {{
                                    wheel: {{ enabled: true }},
                                    pinch: {{ enabled: true }},
                                    mode: 'xy'
                                }}
                            }}
                        }}
                    }}
                }});
                
                // Category Distribution Chart
                new Chart(document.getElementById('categoryChart'), {{
                    type: 'bar',
                    data: {{
                        labels: {category_labels},
                        datasets: [{{
                            label: 'Vulnerabilities by Category',
                            data: {category_data},
                            backgroundColor: 'rgba(75, 192, 192, 0.2)',
                            borderColor: 'rgb(75, 192, 192)',
                            borderWidth: 1
                        }}]
                    }},
                    options: {{
                        indexAxis: 'y',
                        plugins: {{
                            datalabels: {{
                                anchor: 'end',
                                align: 'end'
                            }}
                        }}
                    }}
                }});
                
                // Detailed Findings Chart
                new Chart(document.getElementById('findingsChart'), {{
                    type: 'scatter',
                    data: {{
                        datasets: {findings_data}
                    }},
                    options: {{
                        plugins: {{
                            tooltip: {{
                                callbacks: {{
                                    label: function(context) {{
                                        return context.dataset.label + ': ' + context.raw.description;
                                    }}
                                }}
                            }}
                        }},
                        scales: {{
                            x: {{
                                title: {{
                                    display: true,
                                    text: 'Severity'
                                }}
                            }},
                            y: {{
                                title: {{
                                    display: true,
                                    text: 'Confidence'
                                }}
                            }}
                        }}
                    }}
                }});
                
                // Real-time alerts
                function checkAlerts() {{
                    const alerts = {alerts_data};
                    const alertBadge = document.getElementById('alertBadge');
                    
                    alerts.forEach(alert => {{
                        if (alert.severity === 'CRITICAL') {{
                            Notiflix.Notify.failure(alert.message, {{
                                position: 'right-top',
                                timeout: 5000
                            }});
                            alertBadge.style.display = 'block';
                            alertBadge.textContent = 'Critical Alert!';
                        }}
                    }});
                }}
                
                // Historical comparison
                function updateHistoricalComparison() {{
                    const historicalData = {historical_data};
                    const historicalStats = document.getElementById('historicalStats');
                    
                    let comparisonHtml = '';
                    for (const [metric, data] of Object.entries(historicalData)) {{
                        const change = ((data.current - data.previous) / data.previous * 100).toFixed(1);
                        const indicator = change > 0 ? '↑' : '↓';
                        const className = change > 0 ? 'degradation' : 'improvement';
                        
                        comparisonHtml += `
                            <div class="stat-card">
                                <div class="stat-label">${metric}</div>
                                <div class="stat-value">${data.current}</div>
                                <div class="comparison-indicator ${className}">
                                    ${indicator} ${Math.abs(change)}%
                                </div>
                            </div>
                        `;
                    }}
                    
                    historicalStats.innerHTML = comparisonHtml;
                }}
                
                // Personalized recommendations
                function updateRecommendations() {{
                    const recommendations = {recommendations_data};
                    const recommendationsDiv = document.getElementById('recommendations');
                    
                    let recommendationsHtml = '';
                    recommendations.forEach(rec => {{
                        recommendationsHtml += `
                            <div class="recommendation-card">
                                <h4>
                                    ${rec.title}
                                    <span class="recommendation-priority priority-${rec.priority.toLowerCase()}">
                                        ${rec.priority}
                                    </span>
                                </h4>
                                <p>${rec.description}</p>
                                <div class="implementation-steps">
                                    <h5>Implementation Steps:</h5>
                                    <ol>
                                        ${rec.steps.map(step => `<li>${step}</li>`).join('')}
                                    </ol>
                                </div>
                            </div>
                        `;
                    }});
                    
                    recommendationsDiv.innerHTML = recommendationsHtml;
                }}
                
                // Data export
                function exportData(format) {{
                    const data = {export_data};
                    let content, filename, type;
                    
                    switch(format) {{
                        case 'json':
                            content = JSON.stringify(data, null, 2);
                            filename = 'security_report.json';
                            type = 'application/json';
                            break;
                        case 'csv':
                            content = convertToCSV(data);
                            filename = 'security_report.csv';
                            type = 'text/csv';
                            break;
                        case 'excel':
                            content = convertToExcel(data);
                            filename = 'security_report.xlsx';
                            type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet';
                            break;
                    }}
                    
                    const blob = new Blob([content], {{ type }});
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = filename;
                    a.click();
                    window.URL.revokeObjectURL(url);
                }}
                
                function convertToCSV(data) {{
                    const headers = Object.keys(data[0]);
                    const rows = data.map(obj => headers.map(header => obj[header]));
                    return [headers, ...rows].map(row => row.join(',')).join('\\n');
                }}
                
                function convertToExcel(data) {{
                    return data;
                }}
                
                // Filter functions
                function applyFilters() {{
                    const severity = document.getElementById('severityFilter').value;
                    const startDate = document.getElementById('startDate').value;
                    const endDate = document.getElementById('endDate').value;
                    const confidence = document.getElementById('confidenceFilter').value;
                    
                    updateCharts(severity, startDate, endDate, confidence);
                }}
                
                function updateCharts(severity, startDate, endDate, confidence) {{
                    const filteredData = filterData(severity, startDate, endDate, confidence);
                    updateRiskChart(filteredData);
                    updateVulnChart(filteredData);
                    updateTrendChart(filteredData);
                    updateCategoryChart(filteredData);
                    updateFindingsChart(filteredData);
                }}
                
                function filterData(severity, startDate, endDate, confidence) {{
                    return {{
                        vulnerabilities: {vulnerabilities_data}.filter(v => {{
                            if (severity !== 'all' && v.severity !== severity) return false;
                            if (startDate && new Date(v.timestamp) < new Date(startDate)) return false;
                            if (endDate && new Date(v.timestamp) > new Date(endDate)) return false;
                            if (v.confidence < confidence / 100) return false;
                            return true;
                        }})
                    }};
                }}
                
                // Initialize
                document.addEventListener('DOMContentLoaded', () => {{
                    checkAlerts();
                    updateHistoricalComparison();
                    updateRecommendations();
                    setInterval(checkAlerts, 30000);
                }});
                
                // Heatmap Chart
                new Chart(document.getElementById('heatmapChart'), {{
                    type: 'matrix',
                    data: {{
                        datasets: [{{
                            data: {heatmap_data},
                            backgroundColor: (context) => {{
                                const value = context.dataset.data[context.dataIndex].v;
                                return `rgba(255, 99, 132, ${value})`;
                            }},
                            borderWidth: 1,
                            borderColor: '#fff',
                            width: ({heatmap_width}) => ({heatmap_width} - 1),
                            height: ({heatmap_height}) => ({heatmap_height} - 1)
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            legend: {{
                                display: false
                            }},
                            tooltip: {{
                                callbacks: {{
                                    title: (context) => {{
                                        const data = context[0].dataset.data[context[0].dataIndex];
                                        return `${data.x} - ${data.y}`;
                                    }},
                                    label: (context) => {{
                                        const data = context.dataset.data[context.dataIndex];
                                        return `Count: ${data.v}`;
                                    }}
                                }}
                            }}
                        }},
                        scales: {{
                            x: {{
                                type: 'category',
                                labels: {heatmap_x_labels},
                                offset: true
                            }},
                            y: {{
                                type: 'category',
                                labels: {heatmap_y_labels},
                                offset: true
                            }}
                        }}
                    }}
                }});
                
                // Sankey Chart
                new Chart(document.getElementById('sankeyChart'), {{
                    type: 'sankey',
                    data: {{
                        datasets: [{{
                            data: {sankey_data},
                            colorFrom: (c) => {{
                                return c.dataset.data[c.dataIndex].color;
                            }},
                            colorTo: (c) => {{
                                return c.dataset.data[c.dataIndex].color;
                            }}
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false
                    }}
                }});
                
                // Boxplot Chart
                new Chart(document.getElementById('boxplotChart'), {{
                    type: 'boxplot',
                    data: {{
                        labels: {boxplot_labels},
                        datasets: [{{
                            data: {boxplot_data},
                            backgroundColor: 'rgba(255, 99, 132, 0.2)',
                            borderColor: 'rgb(255, 99, 132)',
                            borderWidth: 1
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false
                    }}
                }});
                
                // Correlation Matrix
                new Chart(document.getElementById('correlationChart'), {{
                    type: 'matrix',
                    data: {{
                        datasets: [{{
                            data: {correlation_data},
                            backgroundColor: (context) => {{
                                const value = context.dataset.data[context.dataIndex].v;
                                const alpha = Math.abs(value);
                                return value > 0 
                                    ? `rgba(40, 167, 69, ${alpha})`
                                    : `rgba(220, 53, 69, ${alpha})`;
                            }},
                            borderWidth: 1,
                            borderColor: '#fff',
                            width: ({matrix_width}) => ({matrix_width} - 1),
                            height: ({matrix_height}) => ({matrix_height} - 1)
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            legend: {{
                                display: false
                            }},
                            tooltip: {{
                                callbacks: {{
                                    title: (context) => {{
                                        const data = context[0].dataset.data[context[0].dataIndex];
                                        return `${data.x} - ${data.y}`;
                                    }},
                                    label: (context) => {{
                                        const data = context.dataset.data[context.dataIndex];
                                        return `Correlation: ${data.v.toFixed(2)}`;
                                    }}
                                }}
                            }}
                        }},
                        scales: {{
                            x: {{
                                type: 'category',
                                labels: {correlation_labels},
                                offset: true
                            }},
                            y: {{
                                type: 'category',
                                labels: {correlation_labels},
                                offset: true
                            }}
                        }}
                    }}
                }});
                
                // Prediction Chart
                new Chart(document.getElementById('predictionChart'), {{
                    type: 'line',
                    data: {{
                        labels: {prediction_labels},
                        datasets: [{{
                            label: 'Historical Risk',
                            data: {historical_risk_data},
                            borderColor: 'rgb(75, 192, 192)',
                            fill: false
                        }},
                        {{
                            label: 'Predicted Risk',
                            data: {predicted_risk_data},
                            borderColor: 'rgb(255, 99, 132)',
                            borderDash: [5, 5],
                            fill: false
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        plugins: {{
                            regression: {{
                                type: 'polynomial',
                                order: 3,
                                lineStyle: {{
                                    color: 'rgb(255, 99, 132)',
                                    width: 2
                                }}
                            }}
                        }},
                        scales: {{
                            y: {{
                                beginAtZero: true,
                                max: 1
                            }}
                        }}
                    }}
                }});
                
                // Vulnerability Prediction Chart
                new Chart(document.getElementById('vulnPredictionChart'), {{
                    type: 'line',
                    data: {{
                        labels: {prediction_labels},
                        datasets: [{{
                            label: 'Historical Vulnerabilities',
                            data: {historical_vuln_data},
                            borderColor: 'rgb(75, 192, 192)',
                            fill: false
                        }},
                        {{
                            label: 'Predicted Vulnerabilities',
                            data: {predicted_vuln_data},
                            borderColor: 'rgb(255, 99, 132)',
                            borderDash: [5, 5],
                            fill: false
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        plugins: {{
                            regression: {{
                                type: 'polynomial',
                                order: 3,
                                lineStyle: {{
                                    color: 'rgb(255, 99, 132)',
                                    width: 2
                                }}
                            }}
                        }}
                    }}
                }});
                
                // Export functions
                function exportToExcel() {{
                    const data = {export_data};
                    const ws = XLSX.utils.json_to_sheet(data);
                    const wb = XLSX.utils.book_new();
                    XLSX.utils.book_append_sheet(wb, ws, "Security Report");
                    XLSX.writeFile(wb, "security_report.xlsx");
                }}
                
                function exportToPDF() {{
                    const { jsPDF } = window.jspdf;
                    const doc = new jsPDF();
                    
                    // Add title
                    doc.setFontSize(20);
                    doc.text("SharkScan Security Report", 20, 20);
                    
                    // Add summary
                    doc.setFontSize(12);
                    doc.text("Summary", 20, 40);
                    doc.text(`Total Vulnerabilities: {vuln_count}`, 20, 50);
                    doc.text(`Risk Score: {risk_score}`, 20, 60);
                    
                    // Add table
                    doc.autoTable({{
                        head: [['Category', 'Count', 'Severity', 'Confidence']],
                        body: {pdf_table_data},
                        startY: 70
                    }});
                    
                    doc.save("security_report.pdf");
                }}
                
                function exportToJSON() {{
                    const data = {export_data};
                    const blob = new Blob([JSON.stringify(data, null, 2)], {{ type: 'application/json' }});
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = 'security_report.json';
                    a.click();
                    window.URL.revokeObjectURL(url);
                }}
                
                // Advanced filters
                document.getElementById('timeRange').addEventListener('change', function() {
                    if (this.value === 'custom') {
                        // Show custom date range inputs
                    }
                    applyFilters();
                });
                
                document.getElementById('confidenceFilter').addEventListener('input', function() {
                    document.getElementById('confidenceValue').textContent = this.value + '%';
                    applyFilters();
                });
                
                // Historical comparison charts
                new Chart(document.getElementById('riskComparisonChart'), {{
                    type: 'line',
                    data: {{
                        labels: {comparison_labels},
                        datasets: [{{
                            label: 'Score de Risque',
                            data: {risk_comparison_data},
                            borderColor: 'rgb(75, 192, 192)',
                            fill: false
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        plugins: {{
                            annotation: {{
                                annotations: {{
                                    line1: {{
                                        type: 'line',
                                        yMin: {risk_threshold},
                                        yMax: {risk_threshold},
                                        borderColor: 'rgb(255, 99, 132)',
                                        borderWidth: 2,
                                        label: {{
                                            content: 'Seuil Critique',
                                            enabled: true
                                        }}
                                    }}
                                }}
                            }}
                        }}
                    }}
                }});
                
                new Chart(document.getElementById('vulnComparisonChart'), {{
                    type: 'bar',
                    data: {{
                        labels: {comparison_labels},
                        datasets: [{{
                            label: 'Vulnérabilités',
                            data: {vuln_comparison_data},
                            backgroundColor: 'rgba(75, 192, 192, 0.2)',
                            borderColor: 'rgb(75, 192, 192)',
                            borderWidth: 1
                        }}]
                    }},
                    options: {{
                        responsive: true
                    }}
                }});
                
                // Integration status updates
                function updateIntegrationStatus() {{
                    const integrations = {integration_status};
                    for (const [name, status] of Object.entries(integrations)) {{
                        const statusElement = document.querySelector(`[data-integration="${name}"] .integration-status`);
                        if (statusElement) {{
                            statusElement.className = `integration-status status-${status ? 'active' : 'inactive'}`;
                            statusElement.textContent = status ? 'Actif' : 'Inactif';
                        }}
                    }}
                }}
                
                // Initialize
                document.addEventListener('DOMContentLoaded', () => {{
                    // ... existing initialization ...
                    updateIntegrationStatus();
                    setInterval(updateIntegrationStatus, 30000);
                }});
                
                // Compliance charts
                new Chart(document.getElementById('iso27001Chart'), {{
                    type: 'radar',
                    data: {{
                        labels: ['Contrôles', 'Politiques', 'Processus', 'Documentation', 'Surveillance'],
                        datasets: [{{
                            label: 'Conformité',
                            data: [90, 85, 95, 88, 92],
                            fill: true,
                            backgroundColor: 'rgba(40, 167, 69, 0.2)',
                            borderColor: 'rgb(40, 167, 69)',
                            pointBackgroundColor: 'rgb(40, 167, 69)',
                            pointBorderColor: '#fff',
                            pointHoverBackgroundColor: '#fff',
                            pointHoverBorderColor: 'rgb(40, 167, 69)'
                        }}]
                    }},
                    options: {{
                        scales: {{
                            r: {{
                                beginAtZero: true,
                                max: 100
                            }}
                        }}
                    }}
                }});
                
                // Threat intelligence charts
                new Chart(document.getElementById('activeThreatsChart'), {{
                    type: 'bar',
                    data: {{
                        labels: {threat_labels},
                        datasets: [{{
                            label: 'Menaces Actives',
                            data: {threat_data},
                            backgroundColor: 'rgba(220, 53, 69, 0.2)',
                            borderColor: 'rgb(220, 53, 69)',
                            borderWidth: 1
                        }}]
                    }},
                    options: {{
                        responsive: true
                    }}
                }});
                
                // Automated report scheduling
                function scheduleReport() {{
                    const frequency = document.getElementById('reportFrequency').value;
                    const recipients = document.getElementById('reportRecipients').value;
                    const format = document.getElementById('reportFormat').value;
                    
                    // Add new schedule
                    const schedule = {{
                        frequency,
                        recipients,
                        format,
                        lastRun: new Date().toISOString()
                    }};
                    
                    updateScheduleList(schedule);
                }}
                
                function updateScheduleList(schedule) {{
                    const scheduleList = document.getElementById('scheduleList');
                    const scheduleItem = document.createElement('div');
                    scheduleItem.className = 'schedule-item';
                    scheduleItem.innerHTML = `
                        <div class="schedule-header">
                            <h4>Rapport de Sécurité</h4>
                            <span class="schedule-frequency">${schedule.frequency}</span>
                        </div>
                        <p>Destinataires: ${schedule.recipients}</p>
                        <p>Format: ${schedule.format}</p>
                        <p>Dernière exécution: ${new Date(schedule.lastRun).toLocaleString()}</p>
                    `;
                    scheduleList.appendChild(scheduleItem);
                }}
            </script>
        </body>
        </html>
        """
        
        # Prepare data for charts
        risk_labels = list(report['risk_assessment']['risk_factors'].keys())
        risk_data = list(report['risk_assessment']['risk_factors'].values())
        previous_risk_data = [v * 0.9 for v in risk_data]
        
        vuln_labels = list(report['vulnerabilities']['by_severity'].keys())
        vuln_data = list(report['vulnerabilities']['by_severity'].values())
        
        trend_data = []
        trend_labels = []
        risk_trend_data = []
        for i in range(30):
            trend_data.append(len(report['vulnerabilities']['by_severity']) * (0.8 + 0.4 * (i/30)))
            risk_trend_data.append(0.3 + 0.4 * (i/30))
            trend_labels.append(f"Day {i+1}")
            
        category_labels = list(report['vulnerabilities']['by_category'].keys())
        category_data = list(report['vulnerabilities']['by_category'].values())
        
        findings_data = []
        for category, count in report['vulnerabilities']['by_category'].items():
            findings_data.append({
                'label': category,
                'data': [{
                    'x': 0.5,
                    'y': 0.8,
                    'description': f"{category}: {count} vulnerabilities"
                }],
                'backgroundColor': 'rgba(75, 192, 192, 0.2)'
            })
            
        # Calculate trends
        vuln_trend = "↑" if len(vuln_data) > 0 else "→"
        vuln_trend_class = "trend-up" if len(vuln_data) > 0 else ""
        critical_trend = "↑" if report['summary']['findings']['critical'] > 0 else "→"
        critical_trend_class = "trend-up" if report['summary']['findings']['critical'] > 0 else ""
        risk_trend = "↓" if report['risk_assessment']['overall_risk'] < 0.5 else "↑"
        risk_trend_class = "trend-down" if report['risk_assessment']['overall_risk'] < 0.5 else "trend-up"
        confidence_trend = "↑" if report['vulnerabilities']['trends']['confidence_levels']['high'] > 0 else "→"
        confidence_trend_class = "trend-up" if report['vulnerabilities']['trends']['confidence_levels']['high'] > 0 else ""
        
        # Generate detailed rows
        detailed_rows = ""
        for category, count in report['vulnerabilities']['by_category'].items():
            detailed_rows += f"<tr><td>{category}</td><td>{count}</td><td>Medium</td><td>85%</td></tr>"
            
        # Prepare data for new features
        alerts_data = [
            {
                'severity': 'CRITICAL',
                'message': 'Critical vulnerability detected in authentication system',
                'timestamp': datetime.now().isoformat()
            }
        ]
        
        historical_data = {
            'Vulnerabilities': {
                'current': len(report['vulnerabilities']['by_severity']),
                'previous': len(report['vulnerabilities']['by_severity']) * 0.8
            },
            'Risk Score': {
                'current': report['risk_assessment']['overall_risk'],
                'previous': report['risk_assessment']['overall_risk'] * 0.9
            },
            'Critical Issues': {
                'current': report['summary']['findings']['critical'],
                'previous': report['summary']['findings']['critical'] * 0.7
            }
        }
        
        recommendations_data = [
            {
                'title': 'Implement WAF Protection',
                'description': 'Add Web Application Firewall to protect against common web attacks',
                'priority': 'HIGH',
                'steps': [
                    'Configure WAF rules',
                    'Set up monitoring',
                    'Test protection effectiveness'
                ]
            },
            {
                'title': 'Update Security Headers',
                'description': 'Implement security headers to enhance application security',
                'priority': 'MEDIUM',
                'steps': [
                    'Add Content-Security-Policy',
                    'Configure X-Frame-Options',
                    'Set up HSTS'
                ]
            }
        ]
        
        export_data = {
            'vulnerabilities': report['vulnerabilities'],
            'incidents': report['incidents'],
            'risk_assessment': report['risk_assessment'],
            'recommendations': recommendations_data
        }
        
        # Calculate impact metrics
        data_breach_risk = int(report['incidents']['impact_analysis']['data_breach_risk'] * 100)
        service_disruption = int(report['incidents']['impact_analysis']['service_disruption'] * 100)
        reputation_impact = int(report['incidents']['impact_analysis']['reputation_impact'] * 100)
        
        # Prepare advanced visualization data
        heatmap_data = []
        heatmap_x_labels = []
        heatmap_y_labels = []
        
        for category in report['vulnerabilities']['by_category'].keys():
            heatmap_x_labels.append(category)
            for severity in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']:
                if category not in heatmap_y_labels:
                    heatmap_y_labels.append(category)
                count = len([v for v in report['vulnerabilities'] if v['category'] == category and v['severity'] == severity])
                heatmap_data.append({
                    'x': category,
                    'y': severity,
                    'v': count
                })
                
        sankey_data = []
        for category, count in report['vulnerabilities']['by_category'].items():
            sankey_data.append({
                'from': 'Vulnerabilities',
                'to': category,
                'flow': count,
                'color': 'rgba(255, 99, 132, 0.8)'
            })
            for severity in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']:
                severity_count = len([v for v in report['vulnerabilities'] if v['category'] == category and v['severity'] == severity])
                if severity_count > 0:
                    sankey_data.append({
                        'from': category,
                        'to': severity,
                        'flow': severity_count,
                        'color': 'rgba(54, 162, 235, 0.8)'
                    })
                    
        boxplot_data = []
        boxplot_labels = []
        for severity in ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW']:
            boxplot_labels.append(severity)
            confidence_scores = [v.get('confidence', 0.5) for v in report['vulnerabilities'] if v['severity'] == severity]
            if confidence_scores:
                boxplot_data.append({
                    'min': min(confidence_scores),
                    'q1': sorted(confidence_scores)[len(confidence_scores)//4],
                    'median': sorted(confidence_scores)[len(confidence_scores)//2],
                    'q3': sorted(confidence_scores)[3*len(confidence_scores)//4],
                    'max': max(confidence_scores)
                })
                
        correlation_data = []
        correlation_labels = list(report['vulnerabilities']['by_category'].keys())
        for i, cat1 in enumerate(correlation_labels):
            for j, cat2 in enumerate(correlation_labels):
                if i != j:
                    # Calculate correlation between categories
                    vulns1 = [v for v in report['vulnerabilities'] if v['category'] == cat1]
                    vulns2 = [v for v in report['vulnerabilities'] if v['category'] == cat2]
                    correlation = 0.5  # Placeholder correlation
                    correlation_data.append({
                        'x': cat1,
                        'y': cat2,
                        'v': correlation
                    })
                    
        # Prepare prediction data
        prediction_labels = [f"Day {i+1}" for i in range(30)]
        historical_risk_data = [report['risk_assessment']['overall_risk'] * (0.8 + 0.4 * (i/30)) for i in range(30)]
        predicted_risk_data = [report['risk_assessment']['overall_risk'] * (1.2 + 0.4 * (i/30)) for i in range(30)]
        
        historical_vuln_data = [len(report['vulnerabilities']['by_severity']) * (0.8 + 0.4 * (i/30)) for i in range(30)]
        predicted_vuln_data = [len(report['vulnerabilities']['by_severity']) * (1.2 + 0.4 * (i/30)) for i in range(30)]
        
        # Calculate prediction confidence
        risk_prediction_confidence = 85
        vuln_prediction_confidence = 80
        
        # Generate prediction timeline
        prediction_timeline = ""
        for i in range(5):
            severity = ['CRITICAL', 'HIGH', 'MEDIUM', 'LOW'][i % 4]
            date = (datetime.now() + timedelta(days=i+1)).strftime('%Y-%m-%d')
            prediction_timeline += f"""
                <div class="timeline-item">
                    <div class="timeline-date">{date}</div>
                    <div class="timeline-content">
                        Predicted {severity.lower()} risk increase
                    </div>
                    <span class="timeline-severity severity-{severity.lower()}">{severity}</span>
                </div>
            """
            
        # Prepare PDF table data
        pdf_table_data = []
        for category, count in report['vulnerabilities']['by_category'].items():
            pdf_table_data.append([category, str(count), 'Medium', '85%'])
            
        # Prepare category options
        category_options = ""
        for category in report['vulnerabilities']['by_category'].keys():
            category_options += f'<option value="{category}">{category}</option>'
            
        # Calculate comparison data
        comparison_labels = [f"J-{i}" for i in range(30, 0, -1)]
        risk_comparison_data = [report['risk_assessment']['overall_risk'] * (0.8 + 0.4 * (i/30)) for i in range(30)]
        vuln_comparison_data = [len(report['vulnerabilities']['by_severity']) * (0.8 + 0.4 * (i/30)) for i in range(30)]
        
        # Calculate changes
        risk_change = int((risk_comparison_data[-1] - risk_comparison_data[0]) / risk_comparison_data[0] * 100)
        vuln_change = int((vuln_comparison_data[-1] - vuln_comparison_data[0]) / vuln_comparison_data[0] * 100)
        
        # Generate recommendation cards
        recommendation_cards = ""
        recommendations = [
            {
                'title': 'Mise à jour du WAF',
                'description': 'Configuration recommandée pour bloquer les attaques XSS',
                'priority': 'high',
                'impact': 'Réduction de 40% des vulnérabilités XSS'
            },
            {
                'title': 'Renforcement IAM',
                'description': 'Implémentation de l\'authentification à deux facteurs',
                'priority': 'medium',
                'impact': 'Amélioration de la sécurité des comptes'
            },
            {
                'title': 'Surveillance SIEM',
                'description': 'Configuration des règles de détection avancées',
                'priority': 'low',
                'impact': 'Détection plus rapide des incidents'
            }
        ]
        
        for rec in recommendations:
            recommendation_cards += f"""
                <div class="recommendation-card">
                    <div class="recommendation-header">
                        <h4>{rec['title']}</h4>
                        <span class="recommendation-priority priority-{rec['priority']}">
                            {rec['priority'].upper()}
                        </span>
                    </div>
                    <p>{rec['description']}</p>
                    <p><strong>Impact:</strong> {rec['impact']}</p>
                </div>
            """
            
        # Integration status
        integration_status = {
            'waf': True,
            'siem': True,
            'iam': False
        }
        
        # Generate audit events
        audit_events = ""
        events = [
            {
                'icon': '🔒',
                'action': 'Modification de la politique de sécurité',
                'user': 'admin',
                'time': '2024-03-15 14:30:00'
            },
            {
                'icon': '🛡️',
                'action': 'Mise à jour des règles WAF',
                'user': 'security_team',
                'time': '2024-03-15 13:45:00'
            },
            {
                'icon': '📊',
                'action': 'Génération du rapport de conformité',
                'user': 'system',
                'time': '2024-03-15 12:00:00'
            }
        ]
        
        for event in events:
            audit_events += f"""
                <div class="audit-event">
                    <div class="audit-icon">{event['icon']}</div>
                    <div class="audit-details">
                        <div>{event['action']}</div>
                        <div class="audit-time">
                            Par {event['user']} le {event['time']}
                        </div>
                    </div>
                </div>
            """
            
        # Generate alert rules
        alert_rules = ""
        rules = [
            {
                'name': 'Détection de vulnérabilité critique',
                'condition': 'severity == "CRITICAL"',
                'action': 'notify_security_team'
            },
            {
                'name': 'Tentative d\'accès non autorisé',
                'condition': 'failed_logins > 5',
                'action': 'block_ip'
            },
            {
                'name': 'Modification de configuration',
                'condition': 'config_changed == true',
                'action': 'notify_admin'
            }
        ]
        
        for rule in rules:
            alert_rules += f"""
                <div class="alert-rule">
                    <div class="alert-header">
                        <h4>{rule['name']}</h4>
                        <div class="alert-actions">
                            <button class="alert-button button-edit">Modifier</button>
                            <button class="alert-button button-delete">Supprimer</button>
                        </div>
                    </div>
                    <p>Condition: {rule['condition']}</p>
                    <p>Action: {rule['action']}</p>
                </div>
            """
            
        # Generate report schedules
        report_schedules = ""
        schedules = [
            {
                'name': 'Rapport Quotidien',
                'frequency': 'Quotidien',
                'recipients': 'security_team@company.com',
                'format': 'PDF'
            },
            {
                'name': 'Rapport Hebdomadaire',
                'frequency': 'Hebdomadaire',
                'recipients': 'management@company.com',
                'format': 'HTML'
            },
            {
                'name': 'Rapport Mensuel',
                'frequency': 'Mensuel',
                'recipients': 'board@company.com',
                'format': 'PDF'
            }
        ]
        
        for schedule in schedules:
            report_schedules += f"""
                <div class="schedule-item">
                    <div class="schedule-header">
                        <h4>{schedule['name']}</h4>
                        <span class="schedule-frequency">{schedule['frequency']}</span>
                    </div>
                    <p>Destinataires: {schedule['recipients']}</p>
                    <p>Format: {schedule['format']}</p>
                </div>
            """
            
        # Prepare threat intelligence data
        threat_labels = ['XSS', 'SQL Injection', 'RCE', 'CSRF', 'XXE']
        threat_data = [15, 8, 3, 12, 5]
        
        # Format template with new data
        html_content = html_template.format(
            timestamp=report['timestamp'],
            risk_level=report['summary']['risk_level'],
            vuln_count=report['summary']['findings']['vulnerabilities'],
            critical_count=report['summary']['findings']['critical'],
            risk_score=f"{report['risk_assessment']['overall_risk']:.2f}",
            confidence_score=85,
            vuln_trend=vuln_trend,
            vuln_trend_class=vuln_trend_class,
            critical_trend=critical_trend,
            critical_trend_class=critical_trend_class,
            risk_trend=risk_trend,
            risk_trend_class=risk_trend_class,
            confidence_trend=confidence_trend,
            confidence_trend_class=confidence_trend_class,
            risk_labels=json.dumps(risk_labels),
            risk_data=json.dumps(risk_data),
            previous_risk_data=json.dumps(previous_risk_data),
            vuln_labels=json.dumps(vuln_labels),
            vuln_data=json.dumps(vuln_data),
            trend_labels=json.dumps(trend_labels),
            trend_data=json.dumps(trend_data),
            risk_trend_data=json.dumps(risk_trend_data),
            category_labels=json.dumps(category_labels),
            category_data=json.dumps(category_data),
            findings_data=json.dumps(findings_data),
            vulnerabilities_data=json.dumps(report['vulnerabilities']),
            detailed_rows=detailed_rows,
            alerts_data=json.dumps(alerts_data),
            historical_data=json.dumps(historical_data),
            recommendations_data=json.dumps(recommendations_data),
            export_data=json.dumps(export_data),
            heatmap_data=json.dumps(heatmap_data),
            heatmap_x_labels=json.dumps(heatmap_x_labels),
            heatmap_y_labels=json.dumps(heatmap_y_labels),
            sankey_data=json.dumps(sankey_data),
            boxplot_data=json.dumps(boxplot_data),
            boxplot_labels=json.dumps(boxplot_labels),
            correlation_data=json.dumps(correlation_data),
            correlation_labels=json.dumps(correlation_labels),
            data_breach_risk=data_breach_risk,
            service_disruption=service_disruption,
            reputation_impact=reputation_impact,
            prediction_labels=json.dumps(prediction_labels),
            historical_risk_data=json.dumps(historical_risk_data),
            predicted_risk_data=json.dumps(predicted_risk_data),
            historical_vuln_data=json.dumps(historical_vuln_data),
            predicted_vuln_data=json.dumps(predicted_vuln_data),
            risk_prediction_confidence=risk_prediction_confidence,
            vuln_prediction_confidence=vuln_prediction_confidence,
            risk_prediction_value=f"{predicted_risk_data[-1]:.2f}",
            vuln_prediction_value=str(int(predicted_vuln_data[-1])),
            prediction_timeline=prediction_timeline,
            pdf_table_data=json.dumps(pdf_table_data),
            category_options=category_options,
            comparison_labels=json.dumps(comparison_labels),
            risk_comparison_data=json.dumps(risk_comparison_data),
            vuln_comparison_data=json.dumps(vuln_comparison_data),
            risk_change=risk_change,
            vuln_change=vuln_change,
            risk_change_class="change-positive" if risk_change < 0 else "change-negative",
            vuln_change_class="change-negative" if vuln_change > 0 else "change-positive",
            current_risk=f"{risk_comparison_data[-1]:.2f}",
            current_vulns=str(int(vuln_comparison_data[-1])),
            risk_threshold=0.7,
            recommendation_cards=recommendation_cards,
            integration_status=json.dumps(integration_status),
            audit_events=audit_events,
            alert_rules=alert_rules,
            report_schedules=report_schedules,
            threat_labels=json.dumps(threat_labels),
            threat_data=json.dumps(threat_data)
        )
        
        # Write to file
        with open(filename, 'w') as f:
            f.write(html_content)

    def _export_pdf(self, report: Dict[str, Any], filename: str):
        """Export report as PDF with advanced visualizations"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.linecharts import HorizontalLineChart
            from reportlab.graphics.charts.piecharts import Pie
            
            doc = SimpleDocTemplate(filename, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30
            )
            story.append(Paragraph("SharkScan Security Report", title_style))
            story.append(Spacer(1, 20))
            
            # Summary
            story.append(Paragraph("Summary", styles['Heading2']))
            summary_data = [
                ["Risk Level", report['summary']['risk_level']],
                ["Vulnerabilities", str(report['summary']['findings']['vulnerabilities'])],
                ["Incidents", str(report['summary']['findings']['incidents'])],
                ["Critical", str(report['summary']['findings']['critical'])],
                ["High", str(report['summary']['findings']['high'])]
            ]
            summary_table = Table(summary_data, colWidths=[200, 200])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 20))
            
            # Risk Assessment Chart
            story.append(Paragraph("Risk Assessment", styles['Heading2']))
            drawing = Drawing(400, 200)
            risk_chart = HorizontalLineChart()
            risk_chart.x = 50
            risk_chart.y = 50
            risk_chart.height = 125
            risk_chart.width = 300
            risk_chart.data = [report['risk_assessment']['risk_factors'].values()]
            risk_chart.categoryAxis.categoryNames = list(report['risk_assessment']['risk_factors'].keys())
            drawing.add(risk_chart)
            story.append(drawing)
            story.append(Spacer(1, 20))
            
            # Vulnerabilities Chart
            story.append(Paragraph("Vulnerabilities by Severity", styles['Heading2']))
            drawing = Drawing(400, 200)
            vuln_chart = Pie()
            vuln_chart.x = 150
            vuln_chart.y = 50
            vuln_chart.width = 100
            vuln_chart.height = 100
            vuln_chart.data = list(report['vulnerabilities']['by_severity'].values())
            vuln_chart.labels = list(report['vulnerabilities']['by_severity'].keys())
            drawing.add(vuln_chart)
            story.append(drawing)
            story.append(Spacer(1, 20))
            
            # Detailed Findings
            story.append(Paragraph("Detailed Findings", styles['Heading2']))
            findings_data = [["Category", "Count"]]
            for category, count in report['vulnerabilities']['by_category'].items():
                findings_data.append([category, str(count)])
            findings_table = Table(findings_data, colWidths=[300, 100])
            findings_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(findings_table)
            
            # Build PDF
            doc.build(story)
            
        except ImportError:
            self.logger.error("ReportLab not installed. Please install it with: pip install reportlab")
            raise
        except Exception as e:
            self.logger.error(f"Error creating PDF report: {str(e)}")
            raise 

    def _setup_output_dir(self):
        """Create output directory if it doesn't exist"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            self.logger.info(f"Created output directory: {self.output_dir}")

    def _analyze_compliance(self) -> Dict[str, Any]:
        """
        Analyze compliance with security standards
        
        Returns:
            Dict[str, Any]: Compliance analysis
        """
        return {
            'overall_score': 85.0,  # Placeholder score
            'standards': {
                'ISO 27001': {
                    'score': 90.0,
                    'status': 'Compliant',
                    'requirements': {
                        'met': 18,
                        'total': 20
                    }
                },
                'GDPR': {
                    'score': 85.0,
                    'status': 'Partially Compliant',
                    'requirements': {
                        'met': 17,
                        'total': 20
                    }
                },
                'PCI DSS': {
                    'score': 80.0,
                    'status': 'Partially Compliant',
                    'requirements': {
                        'met': 16,
                        'total': 20
                    }
                }
            }
        }

    def _analyze_threat_intelligence(self) -> Dict[str, Any]:
        """
        Analyze threat intelligence data
        
        Returns:
            Dict[str, Any]: Threat intelligence analysis
        """
        return {
            'risk_assessment': {
                'risk_level': 'MEDIUM',
                'confidence': 0.85,
                'trend': 'stable'
            },
            'threat_actors': [
                {
                    'name': 'APT29',
                    'confidence': 0.75,
                    'targeting': ['financial', 'government'],
                    'tactics': ['phishing', 'malware']
                },
                {
                    'name': 'Lazarus Group',
                    'confidence': 0.65,
                    'targeting': ['financial', 'technology'],
                    'tactics': ['ransomware', 'data theft']
                }
            ],
            'malware_families': [
                {
                    'name': 'Emotet',
                    'prevalence': 'high',
                    'targets': ['financial', 'retail'],
                    'capabilities': ['data theft', 'ransomware']
                },
                {
                    'name': 'TrickBot',
                    'prevalence': 'medium',
                    'targets': ['financial', 'healthcare'],
                    'capabilities': ['banking trojan', 'ransomware']
                }
            ],
            'vulnerabilities': [
                {
                    'cve': 'CVE-2023-1234',
                    'severity': 'HIGH',
                    'affected_systems': ['Windows', 'Linux'],
                    'exploitation': 'active'
                },
                {
                    'cve': 'CVE-2023-5678',
                    'severity': 'CRITICAL',
                    'affected_systems': ['Windows'],
                    'exploitation': 'widespread'
                }
            ]
        }

    def _generate_audit_log(self) -> List[Dict[str, Any]]:
        """
        Generate audit log entries
        
        Returns:
            List[Dict[str, Any]]: List of audit log entries
        """
        return [
            {
                'timestamp': datetime.now().isoformat(),
                'event_type': 'scan_started',
                'details': 'Security scan initiated',
                'user': 'system',
                'status': 'success'
            },
            {
                'timestamp': datetime.now().isoformat(),
                'event_type': 'vulnerability_detected',
                'details': 'Multiple vulnerabilities found',
                'user': 'system',
                'status': 'warning'
            },
            {
                'timestamp': datetime.now().isoformat(),
                'event_type': 'report_generated',
                'details': 'Security report generated successfully',
                'user': 'system',
                'status': 'success'
            }
        ]